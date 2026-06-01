"""Import data from Accounting_6DE_2026.xlsm into the 6DE platform DB.

Sheets handled:
  - Transactions       (header row 2, data row 3+, columns starting at B)
  - Projects           (header row 3, data row 4+) — project revenue snapshots
  - Recurring Expenses (two stacked sub-tables: yearly ~row 7, monthly ~row 16)
  - CRM                (header row 4, data row 5+) — client YTD revenue merge

Idempotent: uses INSERT OR IGNORE / INSERT OR REPLACE with unique constraints.
"""
from __future__ import annotations

import re
import sys
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Bootstrap
# ---------------------------------------------------------------------------
PLATFORM_ROOT = Path(__file__).resolve().parents[2]
if str(PLATFORM_ROOT) not in sys.path:
    sys.path.insert(0, str(PLATFORM_ROOT))

from db import ensure_db, log_activity  # noqa: E402

import openpyxl  # noqa: E402

# ---------------------------------------------------------------------------
# Source file
# ---------------------------------------------------------------------------
SOURCE = (
    Path(r"C:\Users\Juan\OneDrive - 6th Degree Engineering")
    / "Documents - 6th Degree Engineering"
    / "04_Accounting"
    / "Accounting_6DE_2026.xlsm"
)

# Regex for 6-digit project number (years 22-26)
_JOB_RE = re.compile(r"\b(2[2-6]\d{4})\b")


# ---------------------------------------------------------------------------
# Helpers (shared with import_project_tracker but duplicated for standalone use)
# ---------------------------------------------------------------------------
def _iso(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _float(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _text(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _phone_str(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        digits = str(int(value))
        if len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        return digits
    return str(value).strip() or None


def _lookup_project_id(conn, job_number: str) -> int | None:
    """Look up project id by job_number. Returns None if not found."""
    row = conn.execute(
        "SELECT id FROM projects WHERE job_number = ?", (job_number,)
    ).fetchone()
    return row["id"] if row else None


def _find_header_row(ws, marker: str, max_rows: int = 30) -> int | None:
    """Find the row number whose first non-empty cell contains `marker`."""
    current_row = 0
    for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=False):
        current_row += 1
        for cell in row:
            if cell.value and marker.lower() in str(cell.value).lower():
                return getattr(cell, "row", current_row)
    return None


# ===========================================================================
# IMPORT: Transactions sheet
# ===========================================================================
def import_transactions(conn, wb) -> dict:
    ws = wb["Transactions"]
    stats = {"inserted": 0, "updated": 0, "skipped": 0}

    # Header at row 2, columns starting at B
    headers = []
    for cell in ws[2]:
        headers.append(_text(cell.value))

    col = {}
    for i, h in enumerate(headers):
        if h:
            col[h] = i

    row_num = 2
    for row in ws.iter_rows(min_row=3, values_only=False):
        row_num += 1
        vals = [cell.value for cell in row]
        source_row = getattr(row[0], "row", row_num)

        txn_date = _iso(vals[col.get("Date", 0)])
        if not txn_date:
            stats["skipped"] += 1
            continue

        amount = _float(vals[col.get("Amount", 4)])
        if amount is None:
            stats["skipped"] += 1
            continue

        description = _text(vals[col.get("Transaction Description", 3)])
        account = _text(vals[col.get("Account", 1)])
        account_type = _text(vals[col.get("Account Type", 2)])
        balance = _float(vals[col.get("Balance", 5)])
        expense_category = _text(vals[col.get("Expense Category", 6)])
        txn_type = _text(vals[col.get("Transaction Type", 7)])
        month_val = vals[col.get("Month", 8)]

        # Month: could be an int, float, or string
        month = None
        if month_val is not None:
            try:
                month = int(float(month_val))
            except (ValueError, TypeError):
                # Could be month name
                month_names = {
                    "january": 1, "february": 2, "march": 3, "april": 4,
                    "may": 5, "june": 6, "july": 7, "august": 8,
                    "september": 9, "october": 10, "november": 11, "december": 12,
                }
                if isinstance(month_val, str):
                    month = month_names.get(month_val.strip().lower())

        # Auto-match project by 6-digit number in description
        project_id = None
        if description:
            m = _JOB_RE.search(description)
            if m:
                project_id = _lookup_project_id(conn, m.group(1))

        try:
            conn.execute("""
                INSERT OR IGNORE INTO transactions
                    (txn_date, account, account_type, description, amount,
                     balance, expense_category, txn_type, project_id, month, source_row)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                txn_date, account, account_type, description, amount,
                balance, expense_category, txn_type, project_id, month, source_row,
            ))
            if conn.total_changes:
                stats["inserted"] += 1
            else:
                stats["skipped"] += 1
        except Exception:
            stats["skipped"] += 1

    conn.commit()
    return stats


# ===========================================================================
# IMPORT: Projects sheet (project revenue snapshots)
# ===========================================================================
def import_project_revenue(conn, wb) -> dict:
    ws = wb["Projects"]
    stats = {"inserted": 0, "updated": 0, "skipped": 0}

    # Header at row 3
    headers = []
    for cell in ws[3]:
        headers.append(_text(cell.value))

    col = {}
    for i, h in enumerate(headers):
        if h:
            col[h] = i

    # Find the column for project number — may be labeled "Date/Project No." or similar
    pno_key = None
    for k in col:
        if k and ("project" in k.lower() or "date" in k.lower()):
            pno_key = k
            break

    if pno_key is None:
        print("  WARNING: Could not find project number column in Projects sheet")
        return stats

    service_key = None
    for k in col:
        if k and "service" in k.lower():
            service_key = k
            break

    paid_key = None
    for k in col:
        if k and "amount" in k.lower() and "paid" in k.lower():
            paid_key = k
            break

    cogs_key = None
    for k in col:
        if k and "cogs" in k.lower():
            cogs_key = k
            break

    profit_key = None
    for k in col:
        if k and "profit" in k.lower():
            profit_key = k
            break

    for row in ws.iter_rows(min_row=4, values_only=False):
        vals = [cell.value for cell in row]

        raw_pno = vals[col[pno_key]]
        if raw_pno is None:
            stats["skipped"] += 1
            continue

        # The value is stored as integer (e.g., 231101)
        job_number = str(int(float(raw_pno))) if isinstance(raw_pno, (int, float)) else str(raw_pno).strip()
        if not job_number or not job_number.isdigit():
            stats["skipped"] += 1
            continue
        if len(job_number) < 6:
            job_number = job_number.zfill(6)

        project_id = _lookup_project_id(conn, job_number)
        if not project_id:
            stats["skipped"] += 1
            continue

        service = _text(vals[col[service_key]]) if service_key else None
        amount_paid = _float(vals[col[paid_key]]) if paid_key else None
        cogs = _float(vals[col[cogs_key]]) if cogs_key else None
        profit = _float(vals[col[profit_key]]) if profit_key else None

        # Upsert: delete existing snapshot for this project, insert fresh
        conn.execute("DELETE FROM project_revenue WHERE project_id = ?", (project_id,))
        conn.execute("""
            INSERT INTO project_revenue (project_id, service, amount_paid, cogs, profit)
            VALUES (?, ?, ?, ?, ?)
        """, (project_id, service, amount_paid, cogs, profit))
        stats["inserted"] += 1

    conn.commit()
    return stats


# ===========================================================================
# IMPORT: Recurring Expenses sheet
# ===========================================================================
def import_recurring_expenses(conn, wb) -> dict:
    ws = wb["Recurring Expenses"]
    stats = {"inserted": 0, "updated": 0, "skipped": 0}

    # Read all rows into a list first (read-only mode doesn't support random access)
    all_rows: list[list] = []
    for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
        all_rows.append(list(row))

    # Detect sub-table header rows (0-based in our list)
    yearly_idx = None
    monthly_idx = None
    for i, vals in enumerate(all_rows):
        for v in vals:
            if v is None:
                continue
            vl = str(v).lower()
            if yearly_idx is None and "recurring" in vl and "yearly" in vl:
                yearly_idx = i
            elif monthly_idx is None and "recurring" in vl and "monthly" in vl:
                monthly_idx = i

    def _parse_section(header_idx: int, end_idx: int | None, frequency: str):
        header_vals = all_rows[header_idx]
        headers_lower = [_text(h).lower() if _text(h) else None for h in header_vals]

        # Column indices — data is in cols B+ (index 1+)
        vendor_idx = 1  # column B
        cost_idx = next((i for i, h in enumerate(headers_lower) if h == "cost"), 2)
        due_idx = next((i for i, h in enumerate(headers_lower) if h and "due" in h), None)
        cat_idx = next((i for i, h in enumerate(headers_lower) if h == "category"), None)
        notes_idx = next((i for i, h in enumerate(headers_lower) if h == "notes"), None)

        stop = end_idx if end_idx else min(header_idx + 15, len(all_rows))
        for vals in all_rows[header_idx + 1 : stop]:
            vendor = _text(vals[vendor_idx]) if vendor_idx < len(vals) else None
            if not vendor:
                continue
            if any(kw in vendor.lower() for kw in ("total", "recurring_expenses", "recurring expenses")):
                continue

            cost = _float(vals[cost_idx]) if cost_idx < len(vals) else None
            if cost is None:
                stats["skipped"] += 1
                continue

            monthly_amount = cost / 12.0 if frequency == "annual" else cost
            due_date = _iso(vals[due_idx]) if due_idx is not None and due_idx < len(vals) else None
            category = _text(vals[cat_idx]) if cat_idx is not None and cat_idx < len(vals) else None
            notes_val = _text(vals[notes_idx]) if notes_idx is not None and notes_idx < len(vals) else None

            existing = conn.execute(
                "SELECT id FROM recurring_expenses WHERE LOWER(vendor) = LOWER(?)",
                (vendor,),
            ).fetchone()

            if existing:
                conn.execute("""
                    UPDATE recurring_expenses SET
                        monthly_amount = ?, frequency = ?, next_due_date = ?,
                        category = ?, notes = ?, active = 1
                    WHERE id = ?
                """, (monthly_amount, frequency, due_date, category, notes_val, existing["id"]))
                stats["updated"] += 1
            else:
                conn.execute("""
                    INSERT INTO recurring_expenses (vendor, monthly_amount, frequency, next_due_date, category, notes)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (vendor, monthly_amount, frequency, due_date, category, notes_val))
                stats["inserted"] += 1

    if yearly_idx is not None:
        _parse_section(yearly_idx, monthly_idx, "annual")
    if monthly_idx is not None:
        _parse_section(monthly_idx, None, "monthly")

    if yearly_idx is None and monthly_idx is None:
        print("  WARNING: Could not auto-detect recurring expense sub-tables")

    conn.commit()
    return stats


# ===========================================================================
# IMPORT: CRM sheet (merge client YTD revenue)
# ===========================================================================
def import_crm(conn, wb) -> dict:
    ws = wb["CRM"]
    stats = {"inserted": 0, "updated": 0, "skipped": 0}

    # Header at row 4
    headers = []
    for cell in ws[4]:
        headers.append(_text(cell.value))

    col = {}
    for i, h in enumerate(headers):
        if h:
            col[h] = i

    for row in ws.iter_rows(min_row=5, values_only=False):
        vals = [cell.value for cell in row]

        first_name = _text(vals[col.get("First Name", 1)])
        last_name = _text(vals[col.get("Last Name", 2)])
        if not first_name and not last_name:
            stats["skipped"] += 1
            continue

        full_name = " ".join(filter(None, [first_name, last_name]))
        company = _text(vals[col.get("Company", 3)])
        email = _text(vals[col.get("Email", 4)])
        phone = _phone_str(vals[col.get("Phone", 5)])
        service_type = _text(vals[col.get("Service Type", 6)])
        ytd_revenue = _float(vals[col.get("2026 Account Value", 7)])
        notes = _text(vals[col.get("Notes", 9)])

        # Match by company (case-insensitive)
        existing = None
        if company:
            existing = conn.execute(
                "SELECT id FROM clients WHERE LOWER(company) = LOWER(?)", (company,)
            ).fetchone()
        if not existing:
            existing = conn.execute(
                "SELECT id FROM clients WHERE LOWER(name) = LOWER(?)", (full_name,)
            ).fetchone()

        if existing:
            # Prefer 2026 value for ytd_revenue
            conn.execute("""
                UPDATE clients SET
                    name = ?, company = ?,
                    email = COALESCE(?, email),
                    phone = COALESCE(?, phone),
                    service_type = COALESCE(?, service_type),
                    ytd_revenue = ?,
                    notes = COALESCE(?, notes),
                    updated_at = datetime('now')
                WHERE id = ?
            """, (full_name, company, email, phone, service_type,
                  ytd_revenue, notes, existing["id"]))
            stats["updated"] += 1
        else:
            conn.execute("""
                INSERT INTO clients (name, company, email, phone, service_type, ytd_revenue, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (full_name, company, email, phone, service_type, ytd_revenue, notes))
            stats["inserted"] += 1

    conn.commit()
    return stats


# ===========================================================================
# Main
# ===========================================================================
def main():
    print(f"Source: {SOURCE}")
    if not SOURCE.exists():
        print(f"ERROR: Source file not found: {SOURCE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    conn = ensure_db()


    try:
        # --- Transactions ---
        before = conn.execute("SELECT COUNT(*) FROM transactions").fetchone()[0]
        txn_stats = import_transactions(conn, wb)
        after = conn.execute("SELECT COUNT(*) FROM transactions").fetchone()[0]
        actual_inserted = after - before
        print(f"Transactions:      {actual_inserted} inserted, "
              f"0 updated, {txn_stats['skipped']} skipped")

        # --- Project Revenue ---
        rev_stats = import_project_revenue(conn, wb)
        print(f"Project Revenue:   {rev_stats['inserted']} inserted, "
              f"{rev_stats['updated']} updated, {rev_stats['skipped']} skipped")

        # --- Recurring Expenses ---
        recur_stats = import_recurring_expenses(conn, wb)
        print(f"Recurring Expenses:{recur_stats['inserted']} inserted, "
              f"{recur_stats['updated']} updated, {recur_stats['skipped']} skipped")

        # --- CRM ---
        crm_stats = import_crm(conn, wb)
        print(f"CRM (accounting):  {crm_stats['inserted']} inserted, "
              f"{crm_stats['updated']} updated, {crm_stats['skipped']} skipped")

        # B4/I1: emit a single summary activity_log event per importer run
        log_activity(
            conn,
            entity_type="importer",
            entity_id=0,
            action="imported",
            details={
                "importer": "import_accounting",
                "source": str(SOURCE.name),
                "transactions_inserted": actual_inserted,
                "project_revenue": rev_stats,
                "recurring_expenses": recur_stats,
                "crm": crm_stats,
            },
        )
        conn.commit()
    finally:
        wb.close()
        conn.close()


if __name__ == "__main__":
    main()
