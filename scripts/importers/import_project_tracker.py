"""Import data from Project_Tracker_2026.xlsx into the 6DE platform DB.

Sheets handled:
  - Projects   (header row 3, data row 4+, columns B-V)
  - Proposals  (header row 3, data row 4+)
  - CRM        (header row 2, data row 3+)

Idempotent: uses INSERT OR REPLACE keyed on unique columns.
"""
from __future__ import annotations

import re
import sys
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Bootstrap: add platform root to sys.path so we can import db / config
# ---------------------------------------------------------------------------
PLATFORM_ROOT = Path(__file__).resolve().parents[2]
if str(PLATFORM_ROOT) not in sys.path:
    sys.path.insert(0, str(PLATFORM_ROOT))

from db import ensure_db, log_activity  # noqa: E402

import openpyxl  # noqa: E402

# ---------------------------------------------------------------------------
# Source file
# ---------------------------------------------------------------------------
SOURCE = (
    Path(r"C:\Users\Juan\OneDrive - 6th Degree Engineering")
    / "Documents - 6th Degree Engineering"
    / "06_Engineering"
    / "01_ Active Projects"
    / "Project_Tracker_2026.xlsx"
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _iso(value) -> str | None:
    """Convert an openpyxl datetime cell (or string) to ISO-format date string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # Try common date formats
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _float(value) -> float | None:
    """Safely coerce a cell value to float."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _text(value) -> str | None:
    """Coerce a cell value to stripped string or None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _percent(value) -> float | None:
    """Handle percent values that may be 0-1 floats, 0-100 ints, or '75%' strings."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip().rstrip("%")
        if not value:
            return None
        try:
            v = float(value)
        except ValueError:
            return None
        # If the string had a %, treat as already a percentage
        return v if v > 1 else v * 100
    try:
        v = float(value)
    except (ValueError, TypeError):
        return None
    # openpyxl returns 0.5 for 50%, 1.0 for 100%
    if v <= 1.0:
        return round(v * 100, 2)
    return round(v, 2)


_STATUS_MAP = {
    "active": "active",
    "on hold": "on_hold",
    "on_hold": "on_hold",
    "completed": "completed",
    "complete": "completed",
    "prospect": "prospect",
    "archived": "archived",
    "archive": "archived",
    "closed": "completed",
}


def _normalize_status(raw) -> str:
    """Map various Excel status strings (possibly with emoji) to DB enum values."""
    if raw is None:
        return "active"
    s = str(raw).strip().lower()
    # Strip any non-ASCII characters (emoji)
    s_clean = re.sub(r"[^\x00-\x7F]+", "", s).strip()
    for keyword, status in _STATUS_MAP.items():
        if keyword in s_clean:
            return status
    # Check for common emoji meanings
    if "hold" in s or "⏸" in str(raw) or "pause" in s:
        return "on_hold"
    if "done" in s or "finish" in s or "✅" in str(raw):
        return "completed"
    return "active"


_PROPOSAL_STATUS_MAP = {
    "active": "sent",
    "sent": "sent",
    "pending": "sent",
    "won": "accepted",
    "accepted": "accepted",
    "lost": "declined",
    "declined": "declined",
    "revised": "revised",
    "draft": "draft",
}


def _normalize_proposal_status(raw) -> str:
    if raw is None:
        return "draft"
    s = str(raw).strip().lower()
    s_clean = re.sub(r"[^\x00-\x7F]+", "", s).strip()
    for keyword, status in _PROPOSAL_STATUS_MAP.items():
        if keyword in s_clean:
            return status
    return "draft"


def _phone_str(value) -> str | None:
    """Format a phone number stored as int or float to string."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        digits = str(int(value))
        if len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        return digits
    return str(value).strip() or None


# ---------------------------------------------------------------------------
# Client upsert helper
# ---------------------------------------------------------------------------
def _upsert_client(conn, company: str | None, name: str | None = None,
                   email: str | None = None, phone: str | None = None) -> int | None:
    """Match or create a client row by company name (case-insensitive). Returns client id."""
    if not company and not name:
        return None

    lookup = company or name
    row = conn.execute(
        "SELECT id FROM clients WHERE LOWER(company) = LOWER(?) OR LOWER(name) = LOWER(?)",
        (lookup, lookup),
    ).fetchone()

    if row:
        return row["id"]

    cur = conn.execute(
        "INSERT INTO clients (name, company, email, phone) VALUES (?, ?, ?, ?)",
        (name or company, company, email, phone),
    )
    return cur.lastrowid


# ===========================================================================
# IMPORT: Projects sheet
# ===========================================================================
def import_projects(conn, wb) -> dict:
    ws = wb["Projects"]
    stats = {"inserted": 0, "updated": 0, "skipped": 0}

    # Read header at row 3, columns starting at B (col index 2)
    headers = []
    for cell in ws[3]:
        headers.append(_text(cell.value))

    # Build column-name -> 0-based index map
    col = {}
    for i, h in enumerate(headers):
        if h:
            col[h] = i

    for row in ws.iter_rows(min_row=4, values_only=False):
        vals = [cell.value for cell in row]

        job_number = _text(vals[col.get("Project No", col.get("Project No.", 1))])
        if not job_number:
            stats["skipped"] += 1
            continue

        # Ensure job_number is string
        job_number = str(job_number).strip()
        if job_number.replace(".", "").isdigit():
            job_number = str(int(float(job_number)))
        # Pad to 6 digits if numeric
        if job_number.isdigit() and len(job_number) < 6:
            job_number = job_number.zfill(6)

        name = _text(vals[col.get("Project Description / Address", 2)])
        if not name:
            stats["skipped"] += 1
            continue

        status = _normalize_status(vals[col.get("Project Status", 4)])
        start_date = _iso(vals[col.get("Date Opened", 7)])
        target_end = _iso(vals[col.get("Target Close", 8)])
        company = _text(vals[col.get("Company / Client", 12)])
        contact_name = _text(vals[col.get("Contact", 11)])
        contract_value = _float(vals[col.get("Contract Value ($)", 14)])
        amount_paid = _float(vals[col.get("Amount Paid ($)", 15)])
        outstanding_balance = _float(vals[col.get("Outstanding Balance ($)", 16)])
        cogs = _float(vals[col.get("COGS", 17)])
        profit = _float(vals[col.get("Profit", 18)])
        percent_complete = _percent(vals[col.get("% Complete", 9)])
        priority = _text(vals[col.get("Priority", 3)])
        action_by = _text(vals[col.get("Action By", 5)])
        next_action = _text(vals[col.get("Next Action", 6)])
        folder_path = _text(vals[col.get("Folder", 0)])
        scope = _text(vals[col.get("Scope of Work", 13)])
        notes = _text(vals[col.get("Notes", 20)])

        # Use name as address if it looks like one (contains digits + street-like words)
        address = name if re.search(r"\d+.*(?:st|street|ave|avenue|blvd|dr|rd|way|ct|pl|ln|nw|ne|sw|se)", name, re.IGNORECASE) else None

        client_id = _upsert_client(conn, company, name=contact_name)

        # Check if project exists
        existing = conn.execute(
            "SELECT id FROM projects WHERE job_number = ?", (job_number,)
        ).fetchone()

        if existing:
            conn.execute("""
                UPDATE projects SET
                    name = ?, client_id = ?, address = ?, status = ?,
                    scope = ?, start_date = ?, target_end_date = ?,
                    folder_path = ?, notes = ?, contract_value = ?,
                    amount_paid = ?, outstanding_balance = ?, cogs = ?,
                    profit = ?, percent_complete = ?, priority = ?,
                    action_by = ?, next_action = ?, contact_name = ?,
                    updated_at = datetime('now')
                WHERE job_number = ?
            """, (
                name, client_id, address, status, scope, start_date,
                target_end, folder_path, notes, contract_value,
                amount_paid, outstanding_balance, cogs, profit,
                percent_complete, priority, action_by, next_action,
                contact_name, job_number,
            ))
            stats["updated"] += 1
        else:
            conn.execute("""
                INSERT INTO projects (
                    job_number, name, client_id, address, status, scope,
                    start_date, target_end_date, folder_path, notes,
                    contract_value, amount_paid, outstanding_balance,
                    cogs, profit, percent_complete, priority, action_by,
                    next_action, contact_name
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                job_number, name, client_id, address, status, scope,
                start_date, target_end, folder_path, notes,
                contract_value, amount_paid, outstanding_balance,
                cogs, profit, percent_complete, priority, action_by,
                next_action, contact_name,
            ))
            stats["inserted"] += 1

    conn.commit()
    return stats


# ===========================================================================
# IMPORT: Proposals sheet
# ===========================================================================
def import_proposals(conn, wb) -> dict:
    ws = wb["Proposals"]
    stats = {"inserted": 0, "updated": 0, "skipped": 0}

    headers = []
    for cell in ws[3]:
        headers.append(_text(cell.value))

    col = {}
    for i, h in enumerate(headers):
        if h:
            col[h] = i

    for row in ws.iter_rows(min_row=4, values_only=False):
        vals = [cell.value for cell in row]

        job_number = _text(vals[col.get("Project No", col.get("Project No.", 0))])
        if not job_number:
            stats["skipped"] += 1
            continue

        job_number = str(job_number).strip()
        if job_number.replace(".", "").isdigit():
            job_number = str(int(float(job_number)))
        if job_number.isdigit() and len(job_number) < 6:
            job_number = job_number.zfill(6)

        # Look up project
        proj = conn.execute(
            "SELECT id FROM projects WHERE job_number = ?", (job_number,)
        ).fetchone()

        if not proj:
            # Create a prospect project stub
            proj_name = _text(vals[col.get("Project Description / Address", 1)]) or f"Prospect {job_number}"
            company = _text(vals[col.get("Company / Client", 7)])
            client_id = _upsert_client(conn, company)
            conn.execute("""
                INSERT OR IGNORE INTO projects (job_number, name, client_id, status)
                VALUES (?, ?, ?, 'prospect')
            """, (job_number, proj_name, client_id))
            conn.commit()
            proj = conn.execute(
                "SELECT id FROM projects WHERE job_number = ?", (job_number,)
            ).fetchone()

        project_id = proj["id"]

        proposal_status = _normalize_proposal_status(vals[col.get("Proposal Status", 2)])
        fee_amount = _float(vals[col.get("Contract Value ($)", 10)]) or 0
        scope_text = _text(vals[col.get("Scope of Work", 9)])
        lead_source = _text(vals[col.get("Lead Source", 8)])
        proposal_number = f"P-{job_number}"

        # Update project lead_source if available
        if lead_source:
            conn.execute(
                "UPDATE projects SET lead_source = ? WHERE id = ?",
                (lead_source, project_id),
            )

        # Upsert proposal by proposal_number
        existing = conn.execute(
            "SELECT id FROM proposals WHERE proposal_number = ?", (proposal_number,)
        ).fetchone()

        if existing:
            conn.execute("""
                UPDATE proposals SET
                    project_id = ?, scope_text = ?, fee_amount = ?,
                    status = ?, updated_at = datetime('now')
                WHERE proposal_number = ?
            """, (project_id, scope_text, fee_amount, proposal_status, proposal_number))
            stats["updated"] += 1
        else:
            conn.execute("""
                INSERT INTO proposals (project_id, proposal_number, scope_text, fee_amount, status)
                VALUES (?, ?, ?, ?, ?)
            """, (project_id, proposal_number, scope_text, fee_amount, proposal_status))
            stats["inserted"] += 1

    conn.commit()
    return stats


# ===========================================================================
# IMPORT: CRM sheet
# ===========================================================================
def import_crm(conn, wb) -> dict:
    ws = wb["CRM"]
    stats = {"inserted": 0, "updated": 0, "skipped": 0}

    headers = []
    for cell in ws[2]:
        headers.append(_text(cell.value))

    col = {}
    for i, h in enumerate(headers):
        if h:
            col[h] = i

    for row in ws.iter_rows(min_row=3, values_only=False):
        vals = [cell.value for cell in row]

        first_name = _text(vals[col.get("First Name", 1)])
        last_name = _text(vals[col.get("Last Name", 2)])
        if not first_name and not last_name:
            stats["skipped"] += 1
            continue

        full_name = " ".join(filter(None, [first_name, last_name]))
        company = _text(vals[col.get("Company", 3)])
        email = _text(vals[col.get("Email", 4)])
        phone = _phone_str(vals[col.get("Phone", 5)])
        service_type = _text(vals[col.get("Service Type", 6)])
        ytd_revenue = _float(vals[col.get("2025 Account Value", 7)])
        notes = _text(vals[col.get("Notes", 9)])

        # Match by company (case-insensitive) if company exists, else by name
        existing = None
        if company:
            existing = conn.execute(
                "SELECT id FROM clients WHERE LOWER(company) = LOWER(?)", (company,)
            ).fetchone()
        if not existing:
            existing = conn.execute(
                "SELECT id FROM clients WHERE LOWER(name) = LOWER(?)", (full_name,)
            ).fetchone()

        if existing:
            conn.execute("""
                UPDATE clients SET
                    name = ?, company = ?, email = COALESCE(?, email),
                    phone = COALESCE(?, phone), service_type = ?,
                    ytd_revenue = COALESCE(?, ytd_revenue),
                    notes = COALESCE(?, notes),
                    updated_at = datetime('now')
                WHERE id = ?
            """, (full_name, company, email, phone, service_type,
                  ytd_revenue, notes, existing["id"]))
            stats["updated"] += 1
        else:
            conn.execute("""
                INSERT INTO clients (name, company, email, phone, service_type, ytd_revenue, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (full_name, company, email, phone, service_type, ytd_revenue, notes))
            stats["inserted"] += 1

    conn.commit()
    return stats


# ===========================================================================
# Main
# ===========================================================================
def main():
    print(f"Source: {SOURCE}")
    if not SOURCE.exists():
        print(f"ERROR: Source file not found: {SOURCE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    conn = ensure_db()

    try:
        proj_stats = import_projects(conn, wb)
        print(f"Projects:  {proj_stats['inserted']} inserted, "
              f"{proj_stats['updated']} updated, {proj_stats['skipped']} skipped")

        prop_stats = import_proposals(conn, wb)
        print(f"Proposals: {prop_stats['inserted']} inserted, "
              f"{prop_stats['updated']} updated, {prop_stats['skipped']} skipped")

        crm_stats = import_crm(conn, wb)
        print(f"CRM:       {crm_stats['inserted']} inserted, "
              f"{crm_stats['updated']} updated, {crm_stats['skipped']} skipped")

        # B4/I1: emit a single summary activity_log event per importer run
        log_activity(
            conn,
            entity_type="importer",
            entity_id=0,
            action="imported",
            details={
                "importer": "import_project_tracker",
                "source": str(SOURCE.name),
                "projects": proj_stats,
                "proposals": prop_stats,
                "crm": crm_stats,
            },
        )
        conn.commit()
    finally:
        wb.close()
        conn.close()


if __name__ == "__main__":
    main()
