"""Dry-run / commit importer for legacy xlsx project trackers.

Parses an xlsx sheet, applies a YAML column map, validates each row
through the same workflow rules the platform uses, and emits a per-row
outcome report. Defaults to --dry-run; writing to the database requires
the explicit --commit flag.

Usage examples:
    # Dry-run with default map against the Projects sheet:
    python scripts/import_legacy_xlsx.py --file tracker.xlsx

    # Dry-run with custom map and explicit sheet:
    python scripts/import_legacy_xlsx.py --file tracker.xlsx \\
        --sheet Projects --map config/legacy_map.yaml

    # Write a JSON report:
    python scripts/import_legacy_xlsx.py --file tracker.xlsx \\
        --report import_report.json

    # Actually persist (requires explicit flag):
    python scripts/import_legacy_xlsx.py --file tracker.xlsx --commit
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Bootstrap — add platform root to sys.path
# ---------------------------------------------------------------------------
PLATFORM_ROOT = Path(__file__).resolve().parents[1]
if str(PLATFORM_ROOT) not in sys.path:
    sys.path.insert(0, str(PLATFORM_ROOT))


# ---------------------------------------------------------------------------
# Outcome enum
# ---------------------------------------------------------------------------
class Outcome:
    CREATE = "CREATE"
    UPDATE = "UPDATE-EXISTING"
    SKIP = "SKIP-DUPLICATE"
    FAIL = "FAIL-VALIDATION"


# ---------------------------------------------------------------------------
# Column map loader
# ---------------------------------------------------------------------------
DEFAULT_MAP_PATH = PLATFORM_ROOT / "config" / "legacy_map.yaml"


def load_column_map(path: Path | None = None) -> dict:
    """Load and return the YAML column map as a dict."""
    import yaml  # lazy — only needed when running the script

    map_path = path or DEFAULT_MAP_PATH
    if not map_path.exists():
        print(f"WARNING: Column map not found at {map_path}; using empty map")
        return {"columns": {}, "status_map": {}, "priority_map": {}, "header_row": 1}
    with open(map_path, encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    data.setdefault("columns", {})
    data.setdefault("status_map", {})
    data.setdefault("priority_map", {})
    data.setdefault("header_row", 1)
    return data


# ---------------------------------------------------------------------------
# Cell value helpers (shared with the old importer)
# ---------------------------------------------------------------------------
def _iso(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _text(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _float_val(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Sheet parser
# ---------------------------------------------------------------------------
def parse_sheet(
    file_path: Path,
    sheet_name: str | None = None,
    column_map: dict | None = None,
) -> list[dict[str, Any]]:
    """Parse an xlsx sheet into a list of normalized dicts.

    Each dict's keys are the *platform* field names (post column-map).
    """
    import openpyxl

    col_map = column_map or {}
    columns = col_map.get("columns", {})
    header_row = col_map.get("header_row", 1)

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        if ws is None:
            return []

        # Read header row to build {legacy_header -> col_index} map
        header_cells = list(ws.iter_rows(
            min_row=header_row, max_row=header_row, values_only=True
        ))
        if not header_cells:
            return []
        headers = [_text(c) for c in header_cells[0]]

        # Build {col_index -> platform_field} from the column map
        idx_to_field: dict[int, str] = {}
        for i, hdr in enumerate(headers):
            if hdr is None:
                continue
            # Case-insensitive lookup in the column map
            for legacy_name, platform_field in columns.items():
                if hdr.strip().lower() == legacy_name.strip().lower():
                    idx_to_field[i] = platform_field
                    break

        # Parse data rows
        data_start = header_row + 1
        if "data_start_row" in col_map:
            data_start = col_map["data_start_row"]

        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            record: dict[str, Any] = {}
            for i, val in enumerate(row):
                if i in idx_to_field:
                    record[idx_to_field[i]] = val
            # Skip completely empty rows
            if all(v is None for v in record.values()):
                continue
            rows.append(record)
        return rows
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Row normalizer — apply type coercions and status/priority mapping
# ---------------------------------------------------------------------------
def normalize_row(
    raw: dict[str, Any],
    col_map: dict,
) -> dict[str, Any]:
    """Normalize a parsed row dict into platform-ready values."""
    status_map = col_map.get("status_map", {})
    priority_map = col_map.get("priority_map", {})

    out: dict[str, Any] = {}

    # Job number
    jn = _text(raw.get("job_number"))
    if jn:
        jn = str(jn).strip()
        if jn.replace(".", "").isdigit():
            jn = str(int(float(jn)))
        if jn.isdigit() and len(jn) < 6:
            jn = jn.zfill(6)
    out["job_number"] = jn

    out["name"] = _text(raw.get("name"))

    # Status normalization
    raw_status = _text(raw.get("status"))
    if raw_status:
        s_clean = re.sub(r"[^\x00-\x7F]+", "", raw_status.lower()).strip()
        mapped = None
        for keyword, target in status_map.items():
            if keyword.lower() in s_clean:
                mapped = target
                break
        out["status"] = mapped or "active"
    else:
        out["status"] = "active"

    # Priority normalization
    raw_priority = _text(raw.get("priority"))
    if raw_priority:
        p_clean = re.sub(r"[^\x00-\x7F]+", "", raw_priority.lower()).strip()
        mapped = None
        for keyword, target in priority_map.items():
            if keyword.lower() in p_clean:
                mapped = target
                break
        out["priority"] = mapped
    else:
        out["priority"] = None

    # Dates
    out["start_date"] = _iso(raw.get("start_date"))
    out["target_end_date"] = _iso(raw.get("target_end_date"))

    # Percent complete — use the workflow helper
    from modules.projects.workflow import clamp_percent_complete
    pct_raw = raw.get("percent_complete")
    if pct_raw is not None:
        out["percent_complete"] = clamp_percent_complete(pct_raw)
    else:
        out["percent_complete"] = None

    # Text fields
    for field in ("action_by", "next_action", "scope", "folder_path", "notes"):
        out[field] = _text(raw.get(field))

    # Float fields
    for field in ("contract_value", "amount_paid", "outstanding_balance", "cogs", "profit"):
        out[field] = _float_val(raw.get(field))

    # Client fields (prefixed with _ in the column map)
    out["_client_company"] = _text(raw.get("_client_company"))
    out["_client_contact"] = _text(raw.get("_client_contact"))

    return out


# ---------------------------------------------------------------------------
# Validator — run the same checks the service layer does
# ---------------------------------------------------------------------------
def validate_row(
    row: dict[str, Any],
    row_idx: int,
) -> list[str]:
    """Return a list of validation error strings (empty = valid)."""
    from streamlit_app.components.status_pills import PROJECT_STATUSES
    from modules.projects.workflow import PRIORITY_VALUES

    errors: list[str] = []

    if not row.get("job_number") and not row.get("name"):
        errors.append("Row has neither job_number nor name")

    status = row.get("status", "active")
    if status not in PROJECT_STATUSES:
        errors.append(f"Invalid status: {status!r}")

    priority = row.get("priority")
    if priority is not None and priority not in PRIORITY_VALUES:
        errors.append(f"Invalid priority: {priority!r}")

    return errors


# ---------------------------------------------------------------------------
# Row classifier — determine outcome per row
# ---------------------------------------------------------------------------
def classify_row(
    row: dict[str, Any],
    row_idx: int,
    existing_jobs: set[str],
) -> tuple[str, list[str]]:
    """Return (outcome, errors) for a normalized row."""
    errors = validate_row(row, row_idx)
    if errors:
        return Outcome.FAIL, errors

    jn = row.get("job_number")
    if jn and jn in existing_jobs:
        return Outcome.UPDATE, []

    return Outcome.CREATE, []


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------
def build_report(
    results: list[dict],
    file_path: str,
    sheet_name: str | None,
    commit: bool,
) -> dict:
    """Build a structured report dict from row results."""
    summary = {
        "file": str(file_path),
        "sheet": sheet_name or "(active)",
        "mode": "COMMIT" if commit else "DRY-RUN",
        "total_rows": len(results),
        "create": sum(1 for r in results if r["outcome"] == Outcome.CREATE),
        "update": sum(1 for r in results if r["outcome"] == Outcome.UPDATE),
        "skip": sum(1 for r in results if r["outcome"] == Outcome.SKIP),
        "fail": sum(1 for r in results if r["outcome"] == Outcome.FAIL),
    }
    return {
        "summary": summary,
        "rows": results,
    }


def print_summary(report: dict) -> None:
    """Print a human-readable summary table to stdout."""
    s = report["summary"]
    print()
    print("=" * 60)
    print(f"  Legacy Import Report — {s['mode']}")
    print("=" * 60)
    print(f"  File:     {s['file']}")
    print(f"  Sheet:    {s['sheet']}")
    print(f"  Total:    {s['total_rows']} rows")
    print("-" * 60)
    print(f"  CREATE:          {s['create']}")
    print(f"  UPDATE-EXISTING: {s['update']}")
    print(f"  SKIP-DUPLICATE:  {s['skip']}")
    print(f"  FAIL-VALIDATION: {s['fail']}")
    print("=" * 60)

    # Print failures detail
    failures = [r for r in report["rows"] if r["outcome"] == Outcome.FAIL]
    if failures:
        print()
        print("VALIDATION FAILURES:")
        for f in failures:
            row_num = f["row"]
            jn = f.get("job_number", "?")
            name = f.get("name", "?")
            errs = "; ".join(f["errors"])
            print(f"  Row {row_num}: [{jn}] {name} — {errs}")


def write_report(report: dict, path: Path) -> None:
    """Write the report as JSON."""
    with open(path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)
    print(f"\nReport written to {path}")


# ---------------------------------------------------------------------------
# Commit logic
# ---------------------------------------------------------------------------
def commit_rows(
    results: list[dict],
    normalized_rows: list[dict[str, Any]],
) -> None:
    """Actually persist CREATE / UPDATE rows to the database."""
    from db import ensure_db
    from modules.projects.crud import create_project, update_project
    from modules.projects.codes import next_job_code

    conn = ensure_db()

    for result, row in zip(results, normalized_rows):
        if result["outcome"] == Outcome.FAIL:
            continue

        # Strip internal fields
        data = {k: v for k, v in row.items() if not k.startswith("_") and v is not None}

        if result["outcome"] == Outcome.CREATE:
            if not data.get("job_number"):
                data["job_number"] = next_job_code(conn)
            create_project(conn, **data)
        elif result["outcome"] == Outcome.UPDATE:
            proj = conn.execute(
                "SELECT id FROM projects WHERE job_number = ?",
                (data["job_number"],),
            ).fetchone()
            if proj:
                jn = data.pop("job_number", None)
                data.pop("name", None)  # don't overwrite name on update
                if data:
                    update_project(conn, proj["id"], **data)

    conn.close()


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def run_import(
    file_path: Path,
    sheet_name: str | None = None,
    map_path: Path | None = None,
    commit: bool = False,
    report_path: Path | None = None,
) -> dict:
    """Run the import pipeline and return the report dict."""
    col_map = load_column_map(map_path)
    raw_rows = parse_sheet(file_path, sheet_name=sheet_name, column_map=col_map)

    if not raw_rows:
        print("No data rows found in sheet.")
        return build_report([], str(file_path), sheet_name, commit)

    # Normalize all rows
    normalized = [normalize_row(r, col_map) for r in raw_rows]

    # Get existing job numbers for duplicate detection
    existing_jobs: set[str] = set()
    if commit or True:  # always check for classification
        from db import ensure_db
        conn = ensure_db()
        rows = conn.execute("SELECT job_number FROM projects").fetchall()
        existing_jobs = {r["job_number"] for r in rows}
        conn.close()

    # Classify each row
    results: list[dict] = []
    for i, row in enumerate(normalized):
        outcome, errors = classify_row(row, i, existing_jobs)
        results.append({
            "row": i + 1,
            "job_number": row.get("job_number"),
            "name": row.get("name"),
            "outcome": outcome,
            "errors": errors,
        })

    report = build_report(results, str(file_path), sheet_name, commit)
    print_summary(report)

    if report_path:
        write_report(report, report_path)

    if commit:
        creates = sum(1 for r in results if r["outcome"] == Outcome.CREATE)
        updates = sum(1 for r in results if r["outcome"] == Outcome.UPDATE)
        if creates + updates > 0:
            commit_rows(results, normalized)
            print(f"\nCommitted {creates} creates + {updates} updates to database.")
        else:
            print("\nNothing to commit.")
    else:
        print("\nDRY-RUN mode — no changes written. Use --commit to persist.")

    return report


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import a legacy xlsx project tracker into the 6DE platform."
    )
    parser.add_argument(
        "--file", required=True, type=Path,
        help="Path to the legacy .xlsx file",
    )
    parser.add_argument(
        "--sheet", default=None,
        help="Sheet name (default: first/active sheet)",
    )
    parser.add_argument(
        "--dry-run", action="store_true", default=True, dest="dry_run",
        help="Preview changes without writing (this is the default)",
    )
    parser.add_argument(
        "--commit", action="store_true", default=False,
        help="Actually persist changes to the database",
    )
    parser.add_argument(
        "--map", type=Path, default=None, dest="map_path",
        help="Path to YAML column-map config (default: config/legacy_map.yaml)",
    )
    parser.add_argument(
        "--report", type=Path, default=None, dest="report_path",
        help="Write a structured JSON report to this path",
    )

    args = parser.parse_args()

    if not args.file.exists():
        print(f"ERROR: File not found: {args.file}")
        sys.exit(1)

    run_import(
        file_path=args.file,
        sheet_name=args.sheet,
        map_path=args.map_path,
        commit=args.commit,
        report_path=args.report_path,
    )


if __name__ == "__main__":
    main()
