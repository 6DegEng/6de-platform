"""Dry-run / commit importer for the firm's weekly Excel timesheets.

Scans an HR directory for ``<NN>_<Employee>/02_Timesheets/
Timesheet_<YYYY-MM-DD>_to_<YYYY-MM-DD>.xlsx`` files, parses each file's
``Time_Log`` sheet (headers on row 5, entries from row 6, rows lacking a
Date or Hours skipped — the same rule the Master_Time_Log builder uses)
and imports the rows as platform ``time_entries``.

Mapping rules:
- 6-digit job numbers (e.g. 260304)  -> projects.job_number (missing -> FAIL)
- internal codes (001001..004003)    -> internal_codes.code  (unknown -> FAIL)
- Role display label ("Principal")   -> role enum ("principal")
- OT? = "Y"                          -> multiplier 1.5
- rate = file Rate column / multiplier  (snapshot stays the STD rate, so
  rate * multiplier reproduces the file's Rate column)
- billable = 1 for project rows, 0 for internal rows

Idempotent: a row is skipped when an identical entry already exists
(employee_id, entry_date, project_id/internal_code, hours, description).

DRY-RUN is the default; writing requires the explicit --commit flag.

Usage:
    python scripts/import_timesheets_xlsx.py --hr-dir "<...>/03_Human Resources"
    python scripts/import_timesheets_xlsx.py --hr-dir <dir> --commit
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Bootstrap — add platform root to sys.path
# ---------------------------------------------------------------------------
PLATFORM_ROOT = Path(__file__).resolve().parents[1]
if str(PLATFORM_ROOT) not in sys.path:
    sys.path.insert(0, str(PLATFORM_ROOT))

_FILENAME_RE = re.compile(
    r"^Timesheet_\d{4}-\d{2}-\d{2}_to_\d{4}-\d{2}-\d{2}\.xlsx$", re.IGNORECASE
)
_FOLDER_PREFIX_RE = re.compile(r"^\d+_\s*")


class Outcome:
    CREATE = "CREATE"
    SKIP = "SKIP-DUPLICATE"
    FAIL = "FAIL-VALIDATION"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _norm_code(value: Any) -> str | None:
    """Normalize the Project # cell to a 6-digit string.

    Excel sometimes stores codes as numbers, dropping leading zeros
    (001001 -> 1001), so numeric values are zero-padded back to 6 digits.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        s = str(int(value))
    else:
        s = str(value).strip()
        if not s:
            return None
    if s.isdigit() and len(s) < 6:
        s = s.zfill(6)
    return s


def _norm_name(name: str) -> str:
    return re.sub(r"[^a-z ]", "", name.lower()).strip()


def match_employee(folder_name: str, employees: list) -> Any | None:
    """Match an HR folder name (minus the NN_ prefix) to an employees row.

    Exact case-insensitive match first; otherwise a first+last token match
    (so folder 'Juan Castillo' matches employee 'Juan C. Castillo') — but
    only when exactly one employee matches.
    """
    target = _norm_name(_FOLDER_PREFIX_RE.sub("", folder_name))
    exact = [e for e in employees if _norm_name(e["name"]) == target]
    if len(exact) == 1:
        return exact[0]
    tokens = target.split()
    if len(tokens) >= 2:
        loose = [
            e for e in employees
            if (lambda t: len(t) >= 2 and t[0] == tokens[0] and t[-1] == tokens[-1])
               (_norm_name(e["name"]).split())
        ]
        if len(loose) == 1:
            return loose[0]
    return None


def find_time_log_header_row(ws) -> int | None:
    """Locate the Time_Log header row (expected row 5; scan 1-10 to be safe)."""
    for r in range(1, 11):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if isinstance(a, str) and a.strip() == "Date" \
                and isinstance(b, str) and b.strip() == "Project #":
            return r
    return None


# ---------------------------------------------------------------------------
# Per-file parser
# ---------------------------------------------------------------------------
def parse_time_log(file_path: Path) -> tuple[list[dict], list[str]]:
    """Parse Time_Log rows into dicts. Returns (rows, file_level_errors)."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        if "Time_Log" not in wb.sheetnames:
            return [], ["no Time_Log sheet"]
        ws = wb["Time_Log"]
        # read_only sheets don't support cell() efficiently — materialize.
        grid = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
        header_row = None
        for i, row in enumerate(grid, start=1):
            if row and len(row) >= 2 and \
                    isinstance(row[0], str) and row[0].strip() == "Date" and \
                    isinstance(row[1], str) and row[1].strip() == "Project #":
                header_row = i
                break
        if header_row is None:
            return [], ["Time_Log header row not found (expected row 5)"]

        rows: list[dict] = []
        for idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            row = (tuple(row) + (None,) * 10)[:10]
            (d, code, client, task, role_label, ot, hours, rate,
             line_total, mileage) = row
            entry_date = _iso_date(d)
            if entry_date is None or hours is None:
                continue  # same rule as the Master_Time_Log builder
            rows.append({
                "excel_row": idx,
                "entry_date": entry_date,
                "code": _norm_code(code),
                "client": client,
                "description": str(task).strip() if task is not None else None,
                "role_label": str(role_label).strip() if role_label else None,
                "ot": str(ot).strip().upper() == "Y" if ot else False,
                "hours": float(hours),
                "rate": float(rate) if rate is not None else None,
                "line_total": float(line_total) if line_total is not None else None,
            })
        return rows, []
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Import pipeline
# ---------------------------------------------------------------------------
def run_import(
    hr_dir: Path,
    commit: bool = False,
    report_path: Path | None = None,
) -> dict:
    from db import ensure_db
    from modules.timekeeping.crud import create_time_entry, list_employees
    from modules.timekeeping.export_xlsx import ROLE_FROM_DISPLAY

    conn = ensure_db()
    employees = list_employees(conn)
    projects = {
        r["job_number"]: r["id"]
        for r in conn.execute("SELECT id, job_number FROM projects").fetchall()
    }
    internal = {
        r["code"]: bool(r["is_active"])
        for r in conn.execute("SELECT code, is_active FROM internal_codes").fetchall()
    }

    files = sorted(hr_dir.glob("*/02_Timesheets/Timesheet_*.xlsx"))
    matched_files = [f for f in files if _FILENAME_RE.match(f.name)]
    ignored_files = [f for f in files if not _FILENAME_RE.match(f.name)]

    results: list[dict] = []
    per_employee: dict[str, dict] = {}
    file_summaries: list[dict] = []
    # Mirrors the commit-path duplicate collapse in dry-run mode.
    seen_keys: set[tuple] = set()

    def _dup_exists(key: tuple) -> bool:
        if key in seen_keys:
            return True
        (emp_id, entry_date, project_id, internal_code, hours, desc) = key
        row = conn.execute(
            "SELECT 1 FROM time_entries "
            "WHERE employee_id = ? AND entry_date = ? AND hours = ? "
            "  AND COALESCE(description, '') = ? "
            "  AND COALESCE(project_id, -1) = ? "
            "  AND COALESCE(internal_code, '') = ? "
            "LIMIT 1",
            (emp_id, entry_date, hours, desc or "",
             project_id if project_id is not None else -1,
             internal_code or ""),
        ).fetchone()
        return row is not None

    for f in matched_files:
        folder = f.parents[1].name  # <NN>_<Employee>
        emp = match_employee(folder, employees)
        rows, file_errors = parse_time_log(f)
        fsum = {
            "file": str(f), "folder": folder,
            "employee": emp["name"] if emp else None,
            "rows_parsed": len(rows), "errors": file_errors,
        }
        file_summaries.append(fsum)

        emp_key = emp["name"] if emp else f"UNMATCHED:{folder}"
        stats = per_employee.setdefault(emp_key, {
            "rows": 0, "create": 0, "skip": 0, "fail": 0,
            "hours": 0.0, "dollars": 0.0,
            "file_line_total": 0.0, "computed_line_total": 0.0,
        })

        for r in rows:
            stats["rows"] += 1
            outcome, errors = None, []
            project_id = internal_code = None
            multiplier = 1.5 if r["ot"] else 1.0
            role = ROLE_FROM_DISPLAY.get(r["role_label"] or "")

            if emp is None:
                errors.append(
                    f"unknown employee: folder {folder!r} has no matching "
                    "employees.name"
                )
            code = r["code"]
            if code is None:
                errors.append("missing Project # value")
            elif code in internal:
                if not internal[code]:
                    errors.append(f"internal code {code} is inactive")
                else:
                    internal_code = code
            elif code in projects:
                project_id = projects[code]
            elif code.startswith("00"):
                errors.append(f"unknown internal code: {code}")
            else:
                errors.append(f"job number not found in projects: {code}")

            if role is None:
                errors.append(f"unknown role label: {r['role_label']!r}")
            if r["rate"] is None:
                errors.append("missing Rate value")

            if errors:
                outcome = Outcome.FAIL
                stats["fail"] += 1
            else:
                std_rate = round(r["rate"] / multiplier, 4)
                computed_total = round(r["hours"] * std_rate * multiplier, 2)
                stats["file_line_total"] += r["line_total"] or 0.0
                stats["computed_line_total"] += computed_total
                key = (
                    emp["id"], r["entry_date"], project_id, internal_code,
                    r["hours"], r["description"] or "",
                )
                if _dup_exists(key):
                    outcome = Outcome.SKIP
                    stats["skip"] += 1
                else:
                    outcome = Outcome.CREATE
                    stats["create"] += 1
                    stats["hours"] += r["hours"]
                    stats["dollars"] += computed_total
                    seen_keys.add(key)
                    if commit:
                        create_time_entry(
                            conn,
                            employee_id=emp["id"],
                            project_id=project_id,
                            internal_code=internal_code,
                            entry_date=r["entry_date"],
                            hours=r["hours"],
                            role=role,
                            rate=std_rate,
                            multiplier=multiplier,
                            billable=1 if project_id is not None else 0,
                            description=r["description"],
                        )

            results.append({
                "file": f.name, "excel_row": r["excel_row"],
                "employee": emp_key, "date": r["entry_date"],
                "code": r["code"], "hours": r["hours"],
                "outcome": outcome, "errors": errors,
            })

    report = {
        "summary": {
            "hr_dir": str(hr_dir),
            "mode": "COMMIT" if commit else "DRY-RUN",
            "files_found": len(matched_files),
            "files_ignored": [str(f) for f in ignored_files],
            "rows_parsed": len(results),
            "create": sum(1 for r in results if r["outcome"] == Outcome.CREATE),
            "skip_duplicate": sum(1 for r in results if r["outcome"] == Outcome.SKIP),
            "fail": sum(1 for r in results if r["outcome"] == Outcome.FAIL),
        },
        "per_employee": per_employee,
        "files": file_summaries,
        "rows": results,
    }
    _print_summary(report)
    if report_path:
        with open(report_path, "w", encoding="utf-8") as fh:
            json.dump(report, fh, indent=2, default=str)
        print(f"\nReport written to {report_path}")
    if not commit:
        print("\nDRY-RUN mode — no changes written. Use --commit to persist.")
    return report


def _print_summary(report: dict) -> None:
    s = report["summary"]
    print()
    print("=" * 64)
    print(f"  Timesheet Import Report — {s['mode']}")
    print("=" * 64)
    print(f"  HR dir:          {s['hr_dir']}")
    print(f"  Weekly files:    {s['files_found']}"
          + (f"  (ignored: {len(s['files_ignored'])})" if s["files_ignored"] else ""))
    print(f"  Rows parsed:     {s['rows_parsed']}")
    print("-" * 64)
    print(f"  CREATE:          {s['create']}")
    print(f"  SKIP-DUPLICATE:  {s['skip_duplicate']}")
    print(f"  FAIL-VALIDATION: {s['fail']}")
    print("=" * 64)
    for name, st in report["per_employee"].items():
        print(f"\n  {name}: {st['rows']} rows | create {st['create']} "
              f"| skip {st['skip']} | fail {st['fail']}")
        print(f"    hours (create): {st['hours']:.2f}  "
              f"dollars (create): ${st['dollars']:,.2f}")
        print(f"    file Line Total sum:     ${st['file_line_total']:,.2f}")
        print(f"    computed Line Total sum: ${st['computed_line_total']:,.2f}")
    failures = [r for r in report["rows"] if r["outcome"] == Outcome.FAIL]
    if failures:
        print("\nVALIDATION FAILURES:")
        for r in failures:
            print(f"  {r['file']} row {r['excel_row']}: [{r['code']}] "
                  f"{r['date']} {r['hours']}h — {'; '.join(r['errors'])}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import weekly Excel timesheets into the 6DE platform."
    )
    parser.add_argument(
        "--hr-dir", required=True, type=Path,
        help="Path to the 03_Human Resources directory",
    )
    parser.add_argument(
        "--dry-run", action="store_true", default=True, dest="dry_run",
        help="Preview changes without writing (this is the default)",
    )
    parser.add_argument(
        "--commit", action="store_true", default=False,
        help="Actually persist changes to the database",
    )
    parser.add_argument(
        "--report", type=Path, default=None, dest="report_path",
        help="Write a structured JSON report to this path",
    )
    args = parser.parse_args()

    if not args.hr_dir.is_dir():
        print(f"ERROR: Directory not found: {args.hr_dir}")
        sys.exit(1)

    run_import(args.hr_dir, commit=args.commit, report_path=args.report_path)


if __name__ == "__main__":
    main()
