"""Tests for modules.mirror.xlsx — portfolio overview renderer.

Coverage:
- Header row matches the legacy 22-column shape
- Banner row at A1 with bold red font
- Generated sheet exists with metadata rows
- Deterministic output (byte-identical across renders for same input)
- Conditional formatting rules attached (status fills, priority fills,
  data bar, outstanding-completed font rule)
- Empty input renders a valid workbook without crashing
"""

from __future__ import annotations

import io
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

_PLATFORM_ROOT = Path(__file__).resolve().parents[1]
if str(_PLATFORM_ROOT) not in sys.path:
    sys.path.insert(0, str(_PLATFORM_ROOT))

from modules.mirror.xlsx import BANNER_TEXT, HEADERS, render_portfolio_overview  # noqa: E402


FIXED_DATE = date(2026, 5, 23)


def _seed_projects() -> list[dict]:
    return [
        {
            "id": 1, "job_number": "260301", "name": "Brickell Tower",
            "client_name": "Acme", "status": "active", "priority": "high",
            "percent_complete": 42.6, "start_date": "2026-01-15",
            "target_end_date": "2026-09-30", "contract_value": 125000,
            "amount_paid": 50000, "outstanding_balance": 75000,
            "cogs": 30000, "profit": 95000, "action_by": "Juan",
            "next_action": "MEP", "contact_name": "Jane Smith",
            "scope": "Structural", "notes": "Tall building",
        },
        {
            "id": 2, "job_number": "260215", "name": "Coral Gables House",
            "client_name": "Smith Family", "status": "completed", "priority": "normal",
            "percent_complete": 100, "start_date": "2025-12-01",
            "target_end_date": "2026-03-31", "contract_value": 25000,
            "amount_paid": 25000, "outstanding_balance": 0,
            "cogs": 8000, "profit": 17000, "action_by": "Juan",
            "next_action": "", "contact_name": "Mr. Smith",
            "scope": "Pool deck", "notes": "Done",
        },
    ]


# ---------------------------------------------------------------------------
# Structure
# ---------------------------------------------------------------------------
def test_renders_to_bytes():
    out = render_portfolio_overview(_seed_projects(), today=FIXED_DATE)
    assert isinstance(out, bytes)
    assert len(out) > 0


def test_workbook_has_projects_and_generated_sheets():
    out = render_portfolio_overview(_seed_projects(), today=FIXED_DATE)
    wb = load_workbook(io.BytesIO(out))
    assert "Projects" in wb.sheetnames
    assert "Generated" in wb.sheetnames


def test_headers_match_legacy_shape():
    out = render_portfolio_overview(_seed_projects(), today=FIXED_DATE)
    wb = load_workbook(io.BytesIO(out))
    ws = wb["Projects"]
    header_row = [ws.cell(row=2, column=i + 1).value for i in range(len(HEADERS))]
    assert header_row == HEADERS


def test_banner_in_a1():
    out = render_portfolio_overview(_seed_projects(), today=FIXED_DATE)
    wb = load_workbook(io.BytesIO(out))
    ws = wb["Projects"]
    assert ws.cell(row=1, column=1).value == BANNER_TEXT
    assert ws.cell(row=1, column=1).font.bold is True


def test_data_rows_present():
    out = render_portfolio_overview(_seed_projects(), today=FIXED_DATE)
    wb = load_workbook(io.BytesIO(out))
    ws = wb["Projects"]
    # First data row (row 3) — job number should appear in column 2.
    assert ws.cell(row=3, column=2).value == "260301"
    # Second data row.
    assert ws.cell(row=4, column=2).value == "260215"


def test_generated_sheet_records_metadata():
    out = render_portfolio_overview(
        _seed_projects(), today=FIXED_DATE, platform_version="v3.5",
    )
    wb = load_workbook(io.BytesIO(out))
    ws = wb["Generated"]
    rows = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(2, 7)}
    assert rows["Generated on (UTC date)"] == "2026-05-23"
    assert rows["Platform version"] == "v3.5"
    assert rows["Project count"] == 2


# ---------------------------------------------------------------------------
# Idempotency
# ---------------------------------------------------------------------------
def test_byte_identical_across_renders():
    args = dict(projects=_seed_projects(), today=FIXED_DATE, platform_version="v3.5")
    out1 = render_portfolio_overview(**args)
    out2 = render_portfolio_overview(**args)
    assert out1 == out2, "renders must be byte-identical for the same input"


# ---------------------------------------------------------------------------
# Conditional formatting
# ---------------------------------------------------------------------------
def test_conditional_formatting_rules_attached():
    out = render_portfolio_overview(_seed_projects(), today=FIXED_DATE)
    wb = load_workbook(io.BytesIO(out))
    ws = wb["Projects"]
    # At least 2 ranges should have CF rules: % complete data bar + outstanding font.
    assert len(list(ws.conditional_formatting._cf_rules.keys())) >= 2


def test_priority_cell_has_fill():
    out = render_portfolio_overview(_seed_projects(), today=FIXED_DATE)
    wb = load_workbook(io.BytesIO(out))
    ws = wb["Projects"]
    priority_col = HEADERS.index("Priority") + 1
    # First data project has priority='high' → orange fill (F59E0B).
    cell = ws.cell(row=3, column=priority_col)
    assert cell.fill.fgColor.value.upper().endswith("F59E0B")


def test_status_cell_has_fill():
    out = render_portfolio_overview(_seed_projects(), today=FIXED_DATE)
    wb = load_workbook(io.BytesIO(out))
    ws = wb["Projects"]
    status_col = HEADERS.index("Project Status") + 1
    # First data project has status='active' → green fill (1FBA66).
    cell = ws.cell(row=3, column=status_col)
    assert cell.fill.fgColor.value.upper().endswith("1FBA66")


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------
def test_empty_project_list_renders_valid_workbook():
    out = render_portfolio_overview([], today=FIXED_DATE)
    wb = load_workbook(io.BytesIO(out))
    assert "Projects" in wb.sheetnames
    assert "Generated" in wb.sheetnames
    # Header row still present; no data rows.
    ws = wb["Projects"]
    assert ws.cell(row=2, column=1).value == HEADERS[0]
    assert ws.cell(row=3, column=1).value is None


def test_link_column_uses_base_url():
    out = render_portfolio_overview(
        _seed_projects(), today=FIXED_DATE,
        base_url="https://platform.6de.example",
    )
    wb = load_workbook(io.BytesIO(out))
    ws = wb["Projects"]
    link_col = HEADERS.index("Link") + 1
    assert "platform.6de.example" in ws.cell(row=3, column=link_col).value
    assert "260301" in ws.cell(row=3, column=link_col).value


def test_notes_truncated_to_200_chars():
    project = _seed_projects()[0]
    project["notes"] = "x" * 500
    out = render_portfolio_overview([project], today=FIXED_DATE)
    wb = load_workbook(io.BytesIO(out))
    ws = wb["Projects"]
    notes_col = HEADERS.index("Notes") + 1
    value = ws.cell(row=3, column=notes_col).value
    assert len(value) <= 200
