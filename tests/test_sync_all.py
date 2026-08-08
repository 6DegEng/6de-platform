"""The sync must refuse to import numbers it cannot reproduce.

Excel is the source of truth; the platform mirrors it. The one thing worse
than stale financials is confidently WRONG financials, so reconciliation gates
every write: the import only runs when the numbers that will land in the
database provably equal the workbook's own control total.

The bug that motivated most of this (2026-08-08): the accounting import
reported "reconciled OK - net 44,225.13 matches workbook" and then wrote a
database totalling 19,845.51. Two silent-loss paths, both hidden by
INSERT OR IGNORE:
  - CHECK (account_type IN ('Debit','Credit')) rejects the workbook's
    legitimate 'Business' rows        -> 270 rows, $26,798.49
  - UNIQUE (txn_date, amount, description) collapses genuinely repeated
    charges                            -> 11 rows
Reconciliation now models both, so the mismatch is caught BEFORE anything is
written.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest

_PLATFORM_ROOT = Path(__file__).resolve().parents[1]
if str(_PLATFORM_ROOT) not in sys.path:
    sys.path.insert(0, str(_PLATFORM_ROOT))

from scripts import sync_all  # noqa: E402


# ---------------------------------------------------------------------------
# Fake workbook + source plumbing
# ---------------------------------------------------------------------------
class _FakeSource:
    """A Source whose reconciliation and importers are controlled by the test."""

    def __init__(self, tmp_path, *, matches=True, rows=3, raises=False):
        self.key = "tracker"
        self.label = "fake"
        self.control = "test control"
        self.kind = "workbook"   # parsed via openpyxl, like the real workbooks
        self.calls = []
        self._matches = matches
        self._rows = rows
        self._raises = raises
        self.path = tmp_path / "book.xlsx"
        self.path.write_bytes(b"workbook-v1")
        self.resolve = lambda: self.path
        self.importers = [("things", self._import)]

    def reconcile(self, wb):
        return {"importable": self._rows, "skipped": 0, "unreadable_money": 0,
                "contract_total": 100.0, "paid_total": 0.0,
                "matches": self._matches}

    def _import(self, conn, wb):
        self.calls.append("import")
        if self._raises:
            raise RuntimeError("importer blew up")
        return {"inserted": self._rows}

    def prepare(self, payload):
        return payload

    def describe_reconciliation(self, rec):
        return f"{rec['importable']} things"


@pytest.fixture()
def env(tmp_path, monkeypatch):
    """Isolate state file, run log, and workbook loading from the real ones."""
    monkeypatch.setattr(sync_all, "STATE_FILE", tmp_path / "state.json")
    monkeypatch.setattr(sync_all, "RUN_LOG", tmp_path / "runs.jsonl")
    monkeypatch.setattr(sync_all.openpyxl, "load_workbook",
                        lambda *a, **k: _FakeWorkbook())
    monkeypatch.setattr(sync_all, "verify_landed",
                        lambda source, conn, rec: (True, "stubbed"))
    return tmp_path


class _FakeWorkbook:
    def close(self):
        pass


# ---------------------------------------------------------------------------
# The safety property
# ---------------------------------------------------------------------------
def test_mismatch_refuses_to_write_anything(env, db):
    src = _FakeSource(env, matches=False)

    rec = sync_all.sync_source(src, db, commit=True, force=False,
                               force_commit=False, log=lambda *_: None)

    assert rec["status"] == "mismatch"
    assert rec["committed"] is False
    assert src.calls == [], "no importer may run when reconciliation fails"


def test_dry_run_never_writes_even_when_reconciled(env, db):
    src = _FakeSource(env, matches=True)

    rec = sync_all.sync_source(src, db, commit=False, force=False,
                               force_commit=False, log=lambda *_: None)

    assert rec["status"] == "dry-run"
    assert src.calls == []


def test_commit_runs_importers_when_reconciled(env, db):
    src = _FakeSource(env, matches=True)

    rec = sync_all.sync_source(src, db, commit=True, force=False,
                               force_commit=False, log=lambda *_: None)

    assert rec["status"] == "committed"
    assert rec["committed"] is True
    assert src.calls == ["import"]


def test_force_commit_overrides_a_mismatch_but_says_so(env, db):
    """The escape hatch must work AND be impossible to mistake for a clean run."""
    src = _FakeSource(env, matches=False)

    rec = sync_all.sync_source(src, db, commit=True, force=False,
                               force_commit=True, log=lambda *_: None)

    assert rec["status"] == "committed-forced"
    assert rec["forced"] is True
    assert src.calls == ["import"]


def test_importer_failure_rolls_back_and_reports(env, db):
    src = _FakeSource(env, matches=True, raises=True)

    rec = sync_all.sync_source(src, db, commit=True, force=False,
                               force_commit=False, log=lambda *_: None)

    assert rec["status"] == "error"
    assert "importer blew up" in rec["detail"]


# ---------------------------------------------------------------------------
# Hash gate
# ---------------------------------------------------------------------------
def test_unchanged_workbook_is_skipped_on_the_second_run(env, db):
    src = _FakeSource(env, matches=True)
    sync_all.sync_source(src, db, commit=True, force=False,
                         force_commit=False, log=lambda *_: None)
    src.calls.clear()

    rec = sync_all.sync_source(src, db, commit=True, force=False,
                               force_commit=False, log=lambda *_: None)

    assert rec["status"] == "unchanged"
    assert src.calls == []


def test_edited_workbook_is_picked_up(env, db):
    src = _FakeSource(env, matches=True)
    sync_all.sync_source(src, db, commit=True, force=False,
                         force_commit=False, log=lambda *_: None)
    src.path.write_bytes(b"workbook-v2-edited")
    src.calls.clear()

    rec = sync_all.sync_source(src, db, commit=True, force=False,
                               force_commit=False, log=lambda *_: None)

    assert rec["status"] == "committed"
    assert src.calls == ["import"]


def test_dry_run_does_not_advance_the_hash_gate(env, db):
    """Otherwise a dry-run would convince the next run there was nothing to do
    and the data would never actually be imported."""
    src = _FakeSource(env, matches=True)
    sync_all.sync_source(src, db, commit=False, force=False,
                         force_commit=False, log=lambda *_: None)

    rec = sync_all.sync_source(src, db, commit=True, force=False,
                               force_commit=False, log=lambda *_: None)

    assert rec["status"] == "committed"
    assert src.calls == ["import"]


def test_missing_workbook_is_reported_not_raised(env, db):
    src = _FakeSource(env, matches=True)
    src.path.unlink()

    rec = sync_all.sync_source(src, db, commit=True, force=False,
                               force_commit=False, log=lambda *_: None)

    assert rec["status"] == "missing"
    assert src.calls == []


# ---------------------------------------------------------------------------
# Freshness + run log
# ---------------------------------------------------------------------------
def test_freshness_round_trips_through_the_database(db):
    """The sync runs on Juan's PC; the app runs in Azure. Freshness has to live
    in the DB or the app can never show 'data as of'."""
    sync_all.record_freshness(
        db, "tracker",
        {"status": "committed", "at": "2026-08-08T12:00:00Z", "detail": "50 projects"},
    )

    fresh = sync_all.read_freshness(db)

    assert fresh["tracker"]["status"] == "committed"
    assert fresh["tracker"]["at"] == "2026-08-08T12:00:00Z"
    assert fresh["tracker"]["detail"] == "50 projects"


def test_run_log_is_valid_jsonl(env, db):
    src = _FakeSource(env, matches=True)
    rec = sync_all.sync_source(src, db, commit=True, force=False,
                               force_commit=False, log=lambda *_: None)
    sync_all.append_run_log(rec)

    lines = sync_all.RUN_LOG.read_text(encoding="utf-8").strip().splitlines()

    assert len(lines) == 1
    assert json.loads(lines[0])["source"] == "tracker"


def test_corrupt_state_file_does_not_stop_a_sync(env):
    sync_all.STATE_FILE.write_text("{not json", encoding="utf-8")

    assert sync_all.load_state() == {}


# ---------------------------------------------------------------------------
# The reconciliation model itself (the part that caught the $24,379.62 gap)
# ---------------------------------------------------------------------------
def _txn_workbook(rows):
    """Build a minimal Transactions/Cashflow workbook in memory."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append([])  # row 1 spacer — headers live on row 2
    ws.append(["Date", "Account", "Account Type", "Transaction Description",
               "Amount", "Balance"])
    for r in rows:
        ws.append(list(r))
    cf = wb.create_sheet("Cashflow")
    cf.append(["TOTAL", None, None, round(sum(r[4] for r in rows), 2)])
    return wb


def test_reconcile_flags_rows_the_check_constraint_would_reject():
    """'Business' is a legitimate workbook value that schema.sql forbids;
    INSERT OR IGNORE drops those rows without a word."""
    from scripts.importers.import_accounting import reconcile_transactions

    wb = _txn_workbook([
        ("2026-01-01", "A", "Debit", "ok row", 100.0, None),
        ("2026-01-02", "A", "Business", "silently dropped", 250.0, None),
    ])

    rec = reconcile_transactions(wb)

    assert rec["rejected_rows"] == 1
    assert rec["rejected_net"] == 250.0
    assert rec["rejected_account_types"] == {"Business": 1}
    assert rec["matches"] is False, "must refuse: 250.00 would go missing"


def test_reconcile_flags_repeats_the_unique_key_would_collapse():
    """Two identical charges on one day are real; the unique key eats one."""
    from scripts.importers.import_accounting import reconcile_transactions

    wb = _txn_workbook([
        ("2026-01-01", "A", "Debit", "parking", 6.5, None),
        ("2026-01-01", "A", "Debit", "parking", 6.5, None),
    ])

    rec = reconcile_transactions(wb)

    assert rec["collapsed_rows"] == 1
    assert rec["storable_net"] == 6.5
    assert rec["net"] == 13.0
    assert rec["matches"] is False


def test_reconcile_passes_when_every_row_survives_the_write():
    from scripts.importers.import_accounting import reconcile_transactions

    wb = _txn_workbook([
        ("2026-01-01", "A", "Debit", "one", 100.0, None),
        ("2026-01-02", "A", "Credit", "two", -40.0, None),
    ])

    rec = reconcile_transactions(wb)

    assert rec["matches"] is True
    assert rec["storable_net"] == rec["net"] == 60.0


# ---------------------------------------------------------------------------
# CRM bridge (§5 Phase A / ROADMAP §1.1 finding 1)
# ---------------------------------------------------------------------------
def test_sync_creates_opportunities_without_an_app_restart(db):
    """A sync that imports proposals must leave the CRM pipeline populated.

    The bridge otherwise only fires inside ensure_db() at app startup, which
    Streamlit caches per container — so opportunities would not appear until
    Azure happened to restart. That is the "$0 pipeline / 0 opportunities"
    symptom: 91 proposals imported, 0 opportunities, indefinitely.
    """
    from scripts.sync_all import _bridge_opportunities

    db.execute(
        "INSERT INTO projects (job_number, name, status) VALUES ('260101', 'P', 'active')"
    )
    pid = db.execute("SELECT id FROM projects").fetchone()["id"]
    db.execute(
        "INSERT INTO proposals (project_id, proposal_number, status, fee_amount) "
        "VALUES (?, 'PR-1', 'sent', 10000)",
        (pid,),
    )
    db.commit()
    assert db.execute("SELECT COUNT(*) AS n FROM opportunities").fetchone()["n"] == 0

    stats = _bridge_opportunities(db, None)

    assert stats["created"] == 1
    assert db.execute("SELECT COUNT(*) AS n FROM opportunities").fetchone()["n"] == 1


def test_bridge_is_in_the_tracker_sync_path_and_runs_last():
    """Order matters: it can only bridge proposals that already exist."""
    from scripts.sync_all import SOURCES

    names = [name for name, _ in SOURCES["tracker"].importers]

    assert "opportunities" in names, "the bridge is not wired into the sync"
    assert names.index("proposals") < names.index("opportunities")


def test_bridging_twice_creates_nothing_new(db):
    from scripts.sync_all import _bridge_opportunities

    db.execute(
        "INSERT INTO projects (job_number, name, status) VALUES ('260102', 'Q', 'active')"
    )
    pid = db.execute("SELECT id FROM projects").fetchone()["id"]
    db.execute(
        "INSERT INTO proposals (project_id, proposal_number, status, fee_amount) "
        "VALUES (?, 'PR-2', 'sent', 5000)",
        (pid,),
    )
    db.commit()
    _bridge_opportunities(db, None)

    second = _bridge_opportunities(db, None)

    assert second["created"] == 0
    assert db.execute("SELECT COUNT(*) AS n FROM opportunities").fetchone()["n"] == 1
