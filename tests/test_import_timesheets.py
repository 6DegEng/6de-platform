"""End-to-end tests for scripts/import_timesheets_xlsx.py.

Builds a synthetic weekly timesheet in tmp_path using the real Excel
layout (headers row 5, entries from row 6), imports it, and checks
idempotency + dry-run behavior.
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pytest

_PLATFORM_ROOT = Path(__file__).resolve().parents[1]
if str(_PLATFORM_ROOT) not in sys.path:
    sys.path.insert(0, str(_PLATFORM_ROOT))

from scripts.import_timesheets_xlsx import (  # noqa: E402
    Outcome,
    match_employee,
    run_import,
)


@pytest.fixture()
def hr_dir(tmp_path):
    """HR folder with one weekly file for Juan + a TEMPLATE to ignore."""
    import openpyxl

    ts_dir = tmp_path / "01_Juan Castillo" / "02_Timesheets"
    ts_dir.mkdir(parents=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Time_Log"
    ws.cell(row=1, column=1, value="6th Degree Engineering - Billable Time Log")
    ws.cell(row=2, column=1, value="Enter time entries below.")
    headers = ["Date", "Project #", "Client / Project Name", "Task Description",
               "Role", "OT?", "Hours", "Rate ($/hr)", "Line Total ($)", "Mileage"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=5, column=c, value=h)

    rows = [
        # normal billable row
        [datetime(2026, 6, 29), "260304", "Buena Vista", "Recert rework",
         "Professional Engineer", "N", 5, 190, 950, None],
        # OT row: file shows the OT rate (315 x 1.5)
        [datetime(2026, 7, 1), "260304", "Buena Vista", "Site visit OT",
         "Principal", "Y", 2, 472.5, 945, None],
        # internal row — Excel stored the code as a NUMBER (leading zeros lost)
        [datetime(2026, 6, 30), 2002, "Proposal Tracking",
         "000000 - Proposal Tracking", "Principal", "N", 1, 315, 315, None],
        # bad internal code
        [datetime(2026, 7, 2), "000000", "Mystery", "No such code",
         "Principal", "N", 1, 315, 315, None],
        # job number not in projects
        [datetime(2026, 7, 2), "269999", "Ghost Project", "Not tracked",
         "Professional Engineer", "N", 2, 190, 380, None],
        # blank Rate (older files carry uncached VLOOKUP formulas) ->
        # importer derives the std rate from fee_schedule
        [datetime(2026, 7, 3), "260304", "Buena Vista", "Blank rate row",
         "Engineering Technician", "N", 1, None, None, None],
        # row without hours -> ignored by the parser
        [datetime(2026, 7, 3), "260304", "Buena Vista", "placeholder",
         "Professional Engineer", "N", None, None, 0, None],
    ]
    for r, row in enumerate(rows, start=6):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    wb.save(ts_dir / "Timesheet_2026-06-29_to_2026-07-05.xlsx")

    # A template file that must be ignored
    wb2 = openpyxl.Workbook()
    wb2.save(ts_dir / "Timesheet_TEMPLATE.xlsx")
    return tmp_path


def _entries(db):
    return db.execute(
        "SELECT * FROM time_entries ORDER BY entry_date, id"
    ).fetchall()


def _mk_project(db):
    db.execute(
        "INSERT INTO projects (job_number, name) VALUES ('260304', 'Buena Vista')"
    )


def test_match_employee_tolerates_middle_initial(db):
    employees = db.execute("SELECT * FROM employees").fetchall()
    # Seeded employee is 'Juan C. Castillo'; folder says 'Juan Castillo'
    assert match_employee("01_Juan Castillo", employees)["id"] == 1
    assert match_employee("07_Nobody Here", employees) is None


def test_dry_run_writes_nothing(db, hr_dir, monkeypatch):
    _mk_project(db)
    monkeypatch.setattr("db.ensure_db", lambda: db)
    report = run_import(hr_dir, commit=False)
    assert report["summary"]["mode"] == "DRY-RUN"
    assert report["summary"]["files_found"] == 1
    assert len(report["summary"]["files_ignored"]) == 1  # the TEMPLATE
    assert report["summary"]["rows_parsed"] == 6          # no-hours row dropped
    assert report["summary"]["create"] == 4
    assert report["summary"]["fail"] == 2
    assert len(_entries(db)) == 0                          # nothing written


def test_commit_then_rerun_is_idempotent(db, hr_dir, monkeypatch):
    _mk_project(db)
    monkeypatch.setattr("db.ensure_db", lambda: db)

    report = run_import(hr_dir, commit=True)
    assert report["summary"]["create"] == 4
    assert report["summary"]["skip_duplicate"] == 0
    assert report["summary"]["fail"] == 2

    entries = _entries(db)
    assert len(entries) == 4
    by_date = {e["entry_date"]: e for e in entries}

    derived = by_date["2026-07-03"]
    assert float(derived["rate"]) == 110.0   # fee_schedule fallback
    assert report["per_employee"]["Juan C. Castillo"]["rates_derived"] == 1

    normal = by_date["2026-06-29"]
    assert normal["project_id"] is not None
    assert float(normal["rate"]) == 190.0
    assert float(normal["multiplier"]) == 1.0
    assert normal["billable"] == 1

    ot = by_date["2026-07-01"]
    assert float(ot["multiplier"]) == 1.5
    assert float(ot["rate"]) == 315.0        # std snapshot, not the OT rate
    assert ot["role"] == "principal"
    # rate x multiplier reproduces the file's Rate column
    assert float(ot["rate"]) * float(ot["multiplier"]) == 472.5

    internal = by_date["2026-06-30"]
    assert internal["internal_code"] == "002002"  # zero-padding restored
    assert internal["project_id"] is None
    assert internal["billable"] == 0

    # failures carry reasons
    fails = [r for r in report["rows"] if r["outcome"] == Outcome.FAIL]
    reasons = "; ".join("; ".join(f["errors"]) for f in fails)
    assert "unknown internal code: 000000" in reasons
    assert "job number not found in projects: 269999" in reasons

    # Re-run: everything is a duplicate, nothing new is written
    report2 = run_import(hr_dir, commit=True)
    assert report2["summary"]["create"] == 0
    assert report2["summary"]["skip_duplicate"] == 4
    assert report2["summary"]["fail"] == 2
    assert len(_entries(db)) == 4


def test_unknown_employee_folder_fails_rows(db, hr_dir, monkeypatch):
    import shutil

    _mk_project(db)
    monkeypatch.setattr("db.ensure_db", lambda: db)
    # Clone Juan's folder as an unknown employee
    src = hr_dir / "01_Juan Castillo"
    shutil.copytree(src, hr_dir / "05_Jane Doe")
    report = run_import(hr_dir, commit=True)
    assert report["summary"]["files_found"] == 2
    unmatched = [
        r for r in report["rows"] if r["employee"].startswith("UNMATCHED:")
    ]
    assert len(unmatched) == 6
    assert all(r["outcome"] == Outcome.FAIL for r in unmatched)
    # Jane's rows were never created
    assert len(_entries(db)) == 4
