"""Importer financial synthesis (feat/importer-financial-synthesis).

The tracker carries Contract Value / Amount Paid / Outstanding Balance per
project but no invoice records, so Financials/AR showed $0 after import.
With --synthesize-financials the importer creates at most two invoices per
project (<job>-L1 paid, <job>-L2 outstanding); with --create-clients it
find-or-creates client records from Company/Contact and links them.

Both flags default OFF — the base import behavior is byte-identical.
All tests run against the throwaway test DB fixture; nothing here can
touch production (commit_rows receives the fixture connection).
"""
from __future__ import annotations

import sys
from pathlib import Path

_PLATFORM_ROOT = Path(__file__).resolve().parents[1]
if str(_PLATFORM_ROOT) not in sys.path:
    sys.path.insert(0, str(_PLATFORM_ROOT))

import pytest  # noqa: E402

from scripts.import_legacy_xlsx import (  # noqa: E402
    Outcome,
    client_identity,
    commit_rows,
    ensure_client,
    run_import,
    synthesize_invoices,
)


# ---------------------------------------------------------------------------
# synthesize_invoices — pure function
# ---------------------------------------------------------------------------
class TestSynthesizeInvoices:
    def test_paid_and_outstanding_make_two_invoices(self):
        row = {
            "job_number": "260101", "start_date": "2026-01-15",
            "amount_paid": 7500.0, "outstanding_balance": 2500.0,
        }
        invs = synthesize_invoices(row)
        assert [i["invoice_number"] for i in invs] == ["260101-L1", "260101-L2"]
        paid, outstanding = invs
        assert paid["status"] == "paid"
        assert paid["amount"] == paid["paid_amount"] == 7500.0
        assert paid["issue_date"] == paid["paid_date"] == "2026-01-15"
        assert outstanding["status"] == "sent"
        assert outstanding["amount"] == 2500.0
        assert outstanding["paid_amount"] == 0

    def test_totals_match_tracker_to_the_penny(self):
        row = {
            "job_number": "260304", "start_date": "2026-03-04",
            "amount_paid": 129066.00, "outstanding_balance": 77890.50,
        }
        invs = synthesize_invoices(row)
        assert sum(i["amount"] for i in invs) == 129066.00 + 77890.50

    def test_zero_dollars_make_no_invoices(self):
        assert synthesize_invoices({"job_number": "260101"}) == []
        assert synthesize_invoices(
            {"job_number": "260101", "amount_paid": 0, "outstanding_balance": 0}
        ) == []

    def test_no_job_number_makes_no_invoices(self):
        assert synthesize_invoices({"amount_paid": 100.0}) == []

    def test_missing_start_date_falls_back_to_today(self):
        from datetime import date
        invs = synthesize_invoices({"job_number": "260101", "amount_paid": 1.0})
        assert invs[0]["issue_date"] == date.today().isoformat()

    def test_notes_say_synthesized(self):
        invs = synthesize_invoices({"job_number": "260101", "amount_paid": 1.0})
        assert "Synthesized" in invs[0]["notes"]


# ---------------------------------------------------------------------------
# client identity / ensure_client
# ---------------------------------------------------------------------------
class TestClients:
    def test_identity_none_when_absent(self):
        assert client_identity({}) is None
        assert client_identity({"_client_company": "  "}) is None

    def test_identity_case_insensitive(self):
        a = client_identity({"_client_company": "ACME Corp", "_client_contact": "Jane"})
        b = client_identity({"_client_company": "acme corp", "_client_contact": "jane"})
        assert a == b

    def test_ensure_client_creates_then_reuses(self, db):
        cid1, created1 = ensure_client(db, "Acme Corp", "Jane Doe")
        cid2, created2 = ensure_client(db, "ACME CORP", "jane doe")
        assert cid1 == cid2
        assert created1 is True and created2 is False
        row = db.execute("SELECT name, company FROM clients WHERE id = ?", (cid1,)).fetchone()
        assert row["name"] == "Jane Doe"
        assert row["company"] == "Acme Corp"
        assert db.execute("SELECT COUNT(*) AS c FROM clients").fetchone()["c"] == 1

    def test_ensure_client_company_only(self, db):
        cid, created = ensure_client(db, "Solo LLC", None)
        assert created is True
        row = db.execute("SELECT name FROM clients WHERE id = ?", (cid,)).fetchone()
        assert row["name"] == "Solo LLC"

    def test_ensure_client_nothing_returns_none(self, db):
        assert ensure_client(db, None, None) is None
        assert ensure_client(db, " ", "") is None

    def test_whitespace_variants_dedupe_to_one_client(self, db):
        # Internal whitespace is collapsed for identity — one client only.
        cid1, created1 = ensure_client(db, "Acme  Corp", "Jane   Doe")
        cid2, created2 = ensure_client(db, " Acme Corp ", "Jane Doe")
        assert cid1 == cid2
        assert created1 is True and created2 is False
        assert db.execute("SELECT COUNT(*) AS c FROM clients").fetchone()["c"] == 1
        # ...and the dry-run identity agrees with ensure_client's key.
        a = client_identity({"_client_company": "Acme  Corp", "_client_contact": "Jane   Doe"})
        b = client_identity({"_client_company": "Acme Corp", "_client_contact": "Jane Doe"})
        assert a == b


# ---------------------------------------------------------------------------
# commit_rows integration (fixture DB only)
# ---------------------------------------------------------------------------
def _result(jn, name, outcome=Outcome.CREATE):
    return {"row": 1, "job_number": jn, "name": name, "outcome": outcome, "errors": []}


def _row(jn, name, **extra):
    return {"job_number": jn, "name": name, "status": "active", **extra}


class TestCommitSynthesis:
    def test_flags_off_creates_nothing_extra(self, db):
        counters = commit_rows(
            [_result("260101", "A")],
            [_row("260101", "A", amount_paid=100.0, _client_company="Acme")],
            conn=db,
        )
        assert counters == {
            "invoices_created": 0, "clients_created": 0, "clients_linked": 0,
            "stale_invoices": 0, "row_errors": [],
        }
        assert db.execute("SELECT COUNT(*) AS c FROM invoices").fetchone()["c"] == 0
        assert db.execute("SELECT COUNT(*) AS c FROM clients").fetchone()["c"] == 0

    def test_synthesis_creates_invoices_and_clients(self, db):
        counters = commit_rows(
            [_result("260101", "A"), _result("260201", "B")],
            [
                _row("260101", "A", start_date="2026-01-15",
                     amount_paid=7500.0, outstanding_balance=2500.0,
                     _client_company="Acme Corp", _client_contact="Jane Doe"),
                _row("260201", "B", outstanding_balance=900.0,
                     _client_company="Acme Corp", _client_contact="Jane Doe"),
            ],
            conn=db,
            synthesize_financials=True,
            create_clients=True,
        )
        assert counters["invoices_created"] == 3  # L1+L2 for A, L2 for B
        assert counters["clients_created"] == 1   # same client deduped
        assert counters["clients_linked"] == 2

        # Dollar totals land where the dashboard reads them.
        outstanding = db.execute(
            "SELECT COALESCE(SUM(amount - paid_amount), 0) AS o "
            "FROM invoices WHERE status IN ('sent', 'overdue')"
        ).fetchone()["o"]
        assert outstanding == 2500.0 + 900.0
        paid = db.execute(
            "SELECT COALESCE(SUM(paid_amount), 0) AS p FROM invoices "
            "WHERE status = 'paid'"
        ).fetchone()["p"]
        assert paid == 7500.0

        # Both projects linked to the ONE client record.
        client_ids = {
            r["client_id"] for r in db.execute(
                "SELECT client_id FROM projects WHERE job_number IN ('260101','260201')"
            ).fetchall()
        }
        assert len(client_ids) == 1 and None not in client_ids

    def test_rerun_is_idempotent(self, db):
        results = [_result("260101", "A")]
        rows = [_row("260101", "A", amount_paid=100.0,
                     _client_company="Acme", _client_contact="Jane")]
        c1 = commit_rows(results, rows, conn=db,
                         synthesize_financials=True, create_clients=True)
        # Second run: project now exists -> UPDATE path; invoices/client must
        # NOT duplicate.
        results2 = [_result("260101", "A", outcome=Outcome.UPDATE)]
        c2 = commit_rows(results2, rows, conn=db,
                         synthesize_financials=True, create_clients=True)
        assert c1["invoices_created"] == 1
        assert c2["invoices_created"] == 0
        assert c2["clients_created"] == 0
        assert db.execute("SELECT COUNT(*) AS c FROM invoices").fetchone()["c"] == 1
        assert db.execute("SELECT COUNT(*) AS c FROM clients").fetchone()["c"] == 1

    def test_existing_client_link_not_overwritten(self, db):
        from modules.projects.crud import create_project
        other, _ = ensure_client(db, "Original LLC", None)
        create_project(db, name="A", job_number="260101", client_id=other)
        commit_rows(
            [_result("260101", "A", outcome=Outcome.UPDATE)],
            [_row("260101", "A", _client_company="Different Corp")],
            conn=db,
            create_clients=True,
        )
        row = db.execute(
            "SELECT client_id FROM projects WHERE job_number = '260101'"
        ).fetchone()
        assert row["client_id"] == other  # only fills NULL, never overwrites

    def test_stale_synthesized_invoice_skipped_not_updated(self, db, capsys):
        # First commit synthesizes 260101-L1 at $100.
        c1 = commit_rows(
            [_result("260101", "A")],
            [_row("260101", "A", amount_paid=100.0)],
            conn=db, synthesize_financials=True,
        )
        assert c1["invoices_created"] == 1
        # Tracker amount changes; re-run must NOT auto-update the invoice.
        c2 = commit_rows(
            [_result("260101", "A", outcome=Outcome.UPDATE)],
            [_row("260101", "A", amount_paid=150.0)],
            conn=db, synthesize_financials=True,
        )
        assert c2["invoices_created"] == 0
        assert c2["stale_invoices"] == 1
        out = capsys.readouterr().out
        assert "SKIPPED-STALE 260101-L1" in out
        amount = db.execute(
            "SELECT amount FROM invoices WHERE invoice_number = '260101-L1'"
        ).fetchone()["amount"]
        assert amount == 100.0  # DB amount unchanged

    def test_hand_edited_invoice_not_flagged_stale(self, db):
        commit_rows(
            [_result("260101", "A")],
            [_row("260101", "A", amount_paid=100.0)],
            conn=db, synthesize_financials=True,
        )
        # Simulate a hand-edited invoice (synth marker removed from notes).
        db.execute(
            "UPDATE invoices SET amount = 999.0, notes = 'manually adjusted' "
            "WHERE invoice_number = '260101-L1'"
        )
        db.commit()
        c2 = commit_rows(
            [_result("260101", "A", outcome=Outcome.UPDATE)],
            [_row("260101", "A", amount_paid=100.0)],
            conn=db, synthesize_financials=True,
        )
        assert c2["stale_invoices"] == 0  # only OUR synthesized rows count

    def test_one_bad_row_does_not_kill_the_run(self, db, capsys):
        from modules.projects.crud import create_project
        # completed -> on_hold is an invalid workflow transition.
        create_project(db, name="Done", job_number="260101", status="completed")
        counters = commit_rows(
            [
                _result("260101", "Done", outcome=Outcome.UPDATE),
                _result("260202", "Good"),
            ],
            [
                _row("260101", "Done", status="on_hold"),
                _row("260202", "Good", amount_paid=50.0),
            ],
            conn=db, synthesize_financials=True,
        )
        # The bad row is recorded and the good row still lands.
        assert len(counters["row_errors"]) == 1
        assert counters["row_errors"][0]["job_number"] == "260101"
        assert "InvalidStatusTransition" in counters["row_errors"][0]["error"]
        assert "ERROR (row" in capsys.readouterr().out
        assert counters["invoices_created"] == 1
        assert db.execute(
            "SELECT COUNT(*) AS c FROM projects WHERE job_number = '260202'"
        ).fetchone()["c"] == 1


# ---------------------------------------------------------------------------
# run_import end-to-end (real xlsx, isolated default DB)
# ---------------------------------------------------------------------------
_TRACKER_HEADERS = [
    "Project No", "Project Description / Address", "Project Status",
    "Date Opened", "Amount Paid ($)", "Outstanding Balance ($)",
    "Company / Client", "Contact",
]


def _make_tracker(path, rows):
    """Write a minimal legacy-tracker xlsx (headers on row 3, per the map)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"
    for ci, header in enumerate(_TRACKER_HEADERS, start=1):
        ws.cell(row=3, column=ci, value=header)
    for ri, row in enumerate(rows, start=4):
        for ci, value in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=value)
    wb.save(path)


@pytest.fixture()
def isolated_default_db(tmp_path, monkeypatch):
    """Point ensure_db()'s default connection at a throwaway DB.

    run_import uses ensure_db() internally (no conn parameter), so redirect
    db.DB_PATH/config.DB_PATH to tmp and clear the cache so nothing here can
    touch the real dev database (sqlite file or pg 'public' schema).
    """
    import config as config_mod

    import db as db_mod

    db_path = tmp_path / "import_e2e.db"
    monkeypatch.setenv("PLATFORM_DB_PATH", str(db_path))
    monkeypatch.setattr(db_mod, "DB_PATH", db_path)
    monkeypatch.setattr(config_mod, "DB_PATH", db_path)
    monkeypatch.setattr(db_mod, "LEGACY_DB_PATH", tmp_path / "no_legacy.db")
    if hasattr(db_mod.ensure_db, "clear"):
        db_mod.ensure_db.clear()
    yield db_mod
    if hasattr(db_mod.ensure_db, "clear"):
        db_mod.ensure_db.clear()


def _table_counts(conn) -> dict[str, int]:
    return {
        t: conn.execute(f"SELECT COUNT(*) AS c FROM {t}").fetchone()["c"]
        for t in ("projects", "invoices", "clients")
    }


class TestRunImportDryRun:
    def test_dry_run_writes_no_rows(self, isolated_default_db, tmp_path):
        _make_tracker(tmp_path / "t.xlsx", [
            ("260901", "Test A", "Active", "2026-01-01", 1000.0, 500.0, "Acme Corp", "Jane"),
            ("260902", "Test B", "Active", None, 250.0, 0.0, "Beta LLC", None),
        ])
        conn = isolated_default_db.ensure_db()  # seeds happen up front
        before = _table_counts(conn)
        report = run_import(
            tmp_path / "t.xlsx", sheet_name="Projects", commit=False,
            synthesize_financials=True, create_clients=True,
        )
        after = _table_counts(conn)
        assert after == before  # dry-run wrote no project/invoice/client rows
        fs = report["summary"]["financial_synthesis"]
        assert fs["invoices_planned"] == 3  # L1+L2 for A, L1 for B
        assert fs["paid_total"] == 1250.0
        assert fs["outstanding_total"] == 500.0
        assert fs["distinct_clients"] == 2

    def test_negative_amount_counted_and_makes_no_invoice(
        self, isolated_default_db, tmp_path, capsys,
    ):
        _make_tracker(tmp_path / "neg.xlsx", [
            ("260903", "Refund Job", "Active", None, -50.0, 200.0, None, None),
        ])
        report = run_import(
            tmp_path / "neg.xlsx", sheet_name="Projects", commit=False,
            synthesize_financials=True,
        )
        fs = report["summary"]["financial_synthesis"]
        assert fs["negative_value_rows"] == 1
        assert fs["negative_value_jobs"] == ["260903"]
        assert fs["invoices_planned"] == 1  # only the L2 outstanding invoice
        assert fs["paid_total"] == 0  # the negative paid produced NO invoice
        assert "NEGATIVE" in capsys.readouterr().out
        # Pure function agrees: negative amounts never make an invoice.
        assert synthesize_invoices(
            {"job_number": "260903", "amount_paid": -50.0}
        ) == []
