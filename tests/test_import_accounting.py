"""Tests for the accounting xlsm importer's dry-run reconciliation.

Focus on ``reconcile_transactions`` — the pure, read-only dry-run that must
match the workbook's Cashflow "Net Cash Flow" TOTAL to the penny before any
``--commit``. Uses a synthetic in-memory workbook (never the real file).
"""
from __future__ import annotations

import datetime

import openpyxl

from scripts.importers.import_accounting import reconcile_transactions, resolve_source


def _build_wb(rows, cashflow_net=None):
    """Build a workbook mirroring the real layout: Transactions header on row 2,
    data from row 3, columns B..J (A is blank). Optional Cashflow TOTAL row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws["A1"] = "#VALUE!"  # junk row 1, like the real file
    header = ["Date", "Account", "Account Type", "Transaction Description",
              "Amount", "Balance", "Expense Category", "Transaction Type", "Month"]
    for i, h in enumerate(header):
        ws.cell(row=2, column=2 + i, value=h)  # start at col B
    for r, row in enumerate(rows, start=3):
        for i, val in enumerate(row):
            ws.cell(row=r, column=2 + i, value=val)
    if cashflow_net is not None:
        cf = wb.create_sheet("Cashflow")
        # TOTAL in col A, Net Cash Flow in col D (index 3)
        cf["A2"] = "TOTAL"
        cf["D2"] = cashflow_net
    return wb


def test_reconcile_sums_amounts_and_matches_cashflow():
    d = datetime.datetime(2026, 1, 2)
    rows = [
        [d, 6485, "Debit", "Client payment 260304", 10000.00, 10000.0, "Income", "Income", 1],
        [d, 6485, "Debit", "Bank fee", -29.95, 9970.05, "Bank Fees", "Expense", 1],
        [d, 6485, "Debit", "Subcontractor", -1500.00, 8470.05, "Subs", "Expense", 1],
    ]
    net = 10000.00 - 29.95 - 1500.00  # 8470.05
    rec = reconcile_transactions(_build_wb(rows, cashflow_net=net))
    assert rec["importable"] == 3
    assert rec["net"] == round(net, 2)
    assert rec["inflow"] == 10000.00
    assert rec["outflow"] == round(-29.95 - 1500.00, 2)
    assert rec["cashflow_net"] == round(net, 2)
    assert rec["matches"] is True


def test_reconcile_skips_rows_without_date_or_amount():
    d = datetime.datetime(2026, 1, 2)
    rows = [
        [d, 6485, "Debit", "Opening balance", None, 4000.89, "Opening Balance", "Balance", 1],  # no amount
        [None, 6485, "Debit", "Orphan", 50.0, None, None, None, None],  # no date
        [d, 6485, "Debit", "Real txn", 200.0, 4200.89, "Income", "Income", 1],
    ]
    rec = reconcile_transactions(_build_wb(rows))
    assert rec["importable"] == 1
    assert rec["skipped"] == 2
    assert rec["net"] == 200.0
    # no Cashflow sheet -> cannot reconcile
    assert rec["cashflow_net"] is None
    assert rec["matches"] is False


def test_reconcile_flags_mismatch():
    d = datetime.datetime(2026, 1, 2)
    rows = [[d, 6485, "Debit", "Txn", 100.0, 100.0, "Income", "Income", 1]]
    rec = reconcile_transactions(_build_wb(rows, cashflow_net=999.99))
    assert rec["net"] == 100.0
    assert rec["matches"] is False


def test_resolve_source_precedence(tmp_path, monkeypatch):
    # CLI file wins
    f = tmp_path / "cli.xlsm"
    assert resolve_source(str(f)) == f
    # env var next
    monkeypatch.setenv("SIXDE_ACCOUNTING_XLSM", str(tmp_path / "env.xlsm"))
    assert resolve_source(None) == tmp_path / "env.xlsm"
    # default uses the current user's home (not a hardcoded C:\Users\Juan)
    monkeypatch.delenv("SIXDE_ACCOUNTING_XLSM", raising=False)
    default = resolve_source(None)
    assert "Accounting_6DE_2026.xlsm" in str(default)
    assert r"\Users\Juan\OneDrive" not in str(default)
