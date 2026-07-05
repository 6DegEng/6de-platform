"""Tests for the weekly timesheet xlsx export (Excel-system parity).

Builds the workbook in memory and reads it back with openpyxl to verify
the exact Time_Log layout the downstream Master_Time_Log builder expects.
"""
from __future__ import annotations

import io
import sys
from datetime import date, datetime
from pathlib import Path

_PLATFORM_ROOT = Path(__file__).resolve().parents[1]
if str(_PLATFORM_ROOT) not in sys.path:
    sys.path.insert(0, str(_PLATFORM_ROOT))

from modules.timekeeping.crud import create_time_entry  # noqa: E402
from modules.timekeeping.export_xlsx import (  # noqa: E402
    ROLE_DISPLAY,
    TIME_LOG_HEADERS,
    build_timesheet_workbook,
    timesheet_filename,
)

MONDAY = date(2026, 6, 29)


def _seed_week(db):
    pid = db.execute(
        "INSERT INTO projects (job_number, name) VALUES ('260304', 'Buena Vista')"
    ).lastrowid
    # Regular billable row
    create_time_entry(
        db, employee_id=1, project_id=pid, entry_date="2026-06-29",
        hours=5, role="professional_engineer", description="Recert rework",
    )
    # OT row (principal 315 -> effective 472.50)
    create_time_entry(
        db, employee_id=1, project_id=pid, entry_date="2026-07-01",
        hours=2, role="principal", multiplier=1.5, description="Site visit OT",
    )
    # Internal row
    create_time_entry(
        db, employee_id=1, internal_code="002002", entry_date="2026-06-30",
        hours=1, role="principal", description="000000 - Proposal Tracking",
    )
    return pid


def _load(db):
    data = build_timesheet_workbook(db, 1, MONDAY)
    import openpyxl
    return openpyxl.load_workbook(io.BytesIO(data))


def test_filename_monday_to_sunday():
    assert timesheet_filename(MONDAY) == "Timesheet_2026-06-29_to_2026-07-05.xlsx"


def test_workbook_sheets_and_headers(db):
    _seed_week(db)
    wb = _load(db)
    assert wb.sheetnames == [
        "Fee_Rates", "Internal_Codes", "Time_Log", "Weekly_Summary"
    ]
    ws = wb["Time_Log"]
    headers = [ws.cell(row=5, column=c).value for c in range(1, 11)]
    assert headers == TIME_LOG_HEADERS
    # rows 3-4 must not carry entries (entries start at row 6)
    assert ws.cell(row=4, column=1).value is None


def test_time_log_values_and_ot_math(db):
    _seed_week(db)
    ws = _load(db)["Time_Log"]
    rows = []
    for r in range(6, 20):
        if ws.cell(row=r, column=1).value is None:
            break
        rows.append([ws.cell(row=r, column=c).value for c in range(1, 11)])
    assert len(rows) == 3

    by_date = {r[0].date().isoformat(): r for r in rows}
    # Date cells are real dates, no formulas anywhere
    assert all(isinstance(r[0], datetime) for r in rows)

    normal = by_date["2026-06-29"]
    assert normal[1] == "260304"                      # Project # = job number
    assert normal[2] == "Buena Vista"                 # Client / Project Name
    assert normal[3] == "Recert rework"
    assert normal[4] == "Professional Engineer"       # display label
    assert normal[5] == "N"
    assert normal[6] == 5
    assert normal[7] == 190                           # std rate
    assert normal[8] == 950                           # 5 x 190

    ot = by_date["2026-07-01"]
    assert ot[4] == "Principal"
    assert ot[5] == "Y"
    assert ot[7] == 472.5                             # 315 x 1.5 (OT rate)
    assert ot[8] == 945                               # 2 x 315 x 1.5

    internal = by_date["2026-06-30"]
    assert internal[1] == "002002"                    # internal code
    assert internal[2] == "Proposals / Quotes / Fee Letters"
    assert internal[5] == "N"
    assert internal[8] == 315


def test_fee_rates_and_internal_codes_sheets(db):
    _seed_week(db)
    wb = _load(db)
    fr = wb["Fee_Rates"]
    assert [fr.cell(row=1, column=c).value for c in (1, 2, 3)] == [
        "Role", "Std Rate ($/hr)", "OT Rate ($/hr)"
    ]
    labels = [fr.cell(row=r, column=1).value for r in range(2, 9)]
    assert labels == list(ROLE_DISPLAY.values())
    assert fr.cell(row=2, column=2).value == 315.0    # Principal std
    assert fr.cell(row=2, column=3).value == 472.5    # Principal OT
    assert fr.cell(row=8, column=2).value == 65.0     # Admin std

    ic = wb["Internal_Codes"]
    assert ic.cell(row=3, column=1).value == "Project #"
    codes = [ic.cell(row=r, column=1).value for r in range(4, 22)]
    assert codes[0] == "001001" and codes[-1] == "004003"
    assert len(codes) == 18


def test_weekly_summary_b2_and_totals(db):
    _seed_week(db)
    ws = _load(db)["Weekly_Summary"]
    b2 = ws.cell(row=2, column=2).value
    assert isinstance(b2, datetime) and b2.date() == MONDAY
    # Role table: header row 5, roles 6-12, TOTAL row 13
    assert ws.cell(row=5, column=1).value == "Role"
    role_rows = {
        ws.cell(row=r, column=1).value: (
            ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value
        )
        for r in range(6, 14)
    }
    assert role_rows["Professional Engineer"] == (5, 950)
    assert role_rows["Principal"] == (3, 1260)        # 945 OT + 315 internal
    assert role_rows["TOTAL"] == (8, 2210)


def test_export_only_includes_selected_week_and_employee(db):
    pid = _seed_week(db)
    # Entry outside the week — must not appear
    create_time_entry(
        db, employee_id=1, project_id=pid, entry_date="2026-07-08",
        hours=3, role="principal",
    )
    ws = _load(db)["Time_Log"]
    dates = []
    for r in range(6, 20):
        v = ws.cell(row=r, column=1).value
        if v is None:
            break
        dates.append(v.date().isoformat())
    assert "2026-07-08" not in dates
    assert len(dates) == 3
