"""Permits are built by walking the project folders, linked by job number.

There is no permits workbook — the folder tree IS the record. An active-project
folder is named ``YYMMDD - Name`` and that prefix is the tracker's Project No,
which is how a permit finds its project.

Most projects have no parsed permit number (1 of 52 on 2026-08-08), so the
register is deliberately two-sourced: numbers scraped from the folders, PLUS a
seeded placeholder for every project the tracker says is in AHJ/Permitting.
Without the seeds the Permits page would show one project and look broken.

The county portal is never contacted — see the lockout note in the module.
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest

_PLATFORM_ROOT = Path(__file__).resolve().parents[1]
if str(_PLATFORM_ROOT) not in sys.path:
    sys.path.insert(0, str(_PLATFORM_ROOT))

from scripts.importers import import_permits as ip  # noqa: E402


@pytest.fixture()
def tree(tmp_path):
    """A miniature active-projects folder."""
    root = tmp_path / "01_Active Projects"
    (root / "260304 - Buena Vista" / "03_Correspondence").mkdir(parents=True)
    (root / "260304 - Buena Vista" / "03_Correspondence"
     / "Submittal_UP26031193.eml").write_text("x", encoding="utf-8")
    (root / "260304 - Buena Vista" / "_CLAUDE_BRIEF.md").write_text(
        "Amendment UPA26050307 filed; roof permit UP26032657 pending.",
        encoding="utf-8",
    )
    (root / "251007 - Brickell D1").mkdir()
    (root / "00_Archive").mkdir()       # not a project folder
    (root / "01_Proposals").mkdir()     # not a project folder
    return root


# ---------------------------------------------------------------------------
# Scanning
# ---------------------------------------------------------------------------
def test_only_job_numbered_folders_are_treated_as_projects(tree):
    scan = ip.scan_projects(tree)

    assert set(scan) == {"260304", "251007"}, "00_Archive/01_Proposals are not projects"


def test_permit_numbers_come_from_filenames_and_brief_text(tree):
    scan = ip.scan_projects(tree)

    found = set(scan["260304"]["permits"])
    assert "UP26031193" in found, "missed the number in the .eml filename"
    assert "UP26032657" in found, "missed the number inside _CLAUDE_BRIEF.md"
    assert "UPA26050307" in found, "missed the UPA amendment form"


def test_projects_without_permits_still_appear(tree):
    scan = ip.scan_projects(tree)

    assert scan["251007"]["permits"] == {}


def test_missing_root_returns_empty_rather_than_raising(tmp_path):
    assert ip.scan_projects(tmp_path / "nope") == {}


def test_reconciliation_fails_on_an_empty_scan(tmp_path):
    """A wrong path would otherwise look like 'no permits' and write nothing
    while reporting success."""
    rec = ip.reconcile_permits(tmp_path / "nope")

    assert rec["matches"] is False


def test_reconciliation_reports_the_real_counts(tree):
    rec = ip.reconcile_permits(tree)

    assert rec["folders"] == 2
    assert rec["with_permits"] == 1
    assert rec["matches"] is True


# ---------------------------------------------------------------------------
# Importing
# ---------------------------------------------------------------------------
def _add_project(conn, job_number, status="active"):
    conn.execute(
        "INSERT INTO projects (job_number, name, status) VALUES (?, ?, ?)",
        (job_number, f"Project {job_number}", status),
    )
    return conn.execute(
        "SELECT id FROM projects WHERE job_number = ?", (job_number,)
    ).fetchone()["id"]


def test_discovered_permits_attach_to_the_right_project(db, tree):
    pid = _add_project(db, "260304")
    _add_project(db, "251007")

    stats = ip.import_permits(db, ip.scan_projects(tree))

    assert stats["inserted"] == 3
    rows = db.execute(
        "SELECT permit_number FROM permits WHERE project_id = ?", (pid,)
    ).fetchall()
    assert {r["permit_number"] for r in rows} == {
        "UP26031193", "UP26032657", "UPA26050307"
    }


def test_running_twice_does_not_duplicate(db, tree):
    _add_project(db, "260304")
    scan = ip.scan_projects(tree)
    ip.import_permits(db, scan)

    second = ip.import_permits(db, scan)

    assert second["inserted"] == 0
    assert second["updated"] == 3
    assert db.execute("SELECT COUNT(*) AS n FROM permits").fetchone()["n"] == 3


def test_folder_with_no_matching_project_is_counted_not_crashed(db, tree):
    """A folder whose job number isn't in the tracker must be reported, not
    silently dropped and not a foreign-key crash."""
    stats = ip.import_permits(db, ip.scan_projects(tree))

    assert stats["no_project"] == 2
    assert stats["inserted"] == 0


def test_ahj_permitting_projects_are_seeded_even_without_a_number(db, tree):
    """The point of seeding: the Permits page should show what is actually in
    permitting, not only what happens to be greppable."""
    _add_project(db, "251007", status="ahj_permitting")

    stats = ip.import_permits(db, ip.scan_projects(tree))

    assert stats["seeded"] == 1
    row = db.execute(
        "SELECT permit_number, status FROM permits"
    ).fetchone()
    assert row["permit_number"] is None
    assert row["status"] == "pending"


def test_seeding_skips_projects_that_already_have_permits(db, tree):
    _add_project(db, "260304", status="ahj_permitting")

    stats = ip.import_permits(db, ip.scan_projects(tree))

    assert stats["seeded"] == 0
    assert stats["skipped"] == 1


def test_permit_type_is_guessed_within_the_schema_s_allowed_values(db, tmp_path):
    """permit_type has a CHECK constraint; an invented value would be rejected
    at write time — the failure mode that cost $24,379.62 in accounting."""
    root = tmp_path / "proj"
    (root / "260101 - Roof Job").mkdir(parents=True)
    (root / "260101 - Roof Job" / "roof_permit_UP26010001.pdf").write_text(
        "x", encoding="utf-8")
    _add_project(db, "260101")

    ip.import_permits(db, ip.scan_projects(root))

    assert db.execute(
        "SELECT permit_type FROM permits"
    ).fetchone()["permit_type"] == "roofing"


# ---------------------------------------------------------------------------
# Orchestration order
# ---------------------------------------------------------------------------
def test_permits_run_after_the_tracker():
    """Permits link to projects by job number, so an alphabetical run order
    imports them before any project exists and drops every one."""
    from scripts.sync_all import SOURCE_ORDER

    assert SOURCE_ORDER.index("tracker") < SOURCE_ORDER.index("permits")


# ---------------------------------------------------------------------------
# §1.1 finding 5 — Completed projects must not show 0% progress
# ---------------------------------------------------------------------------
def test_completed_projects_are_forced_to_100_percent(db, tmp_path):
    """The tracker's "% Complete" column was never backfilled when jobs closed,
    so finished projects read "Completed / 1%" and made the whole column
    untrustworthy. Status is the reliable signal."""
    import openpyxl

    from scripts.importers import import_project_tracker as trk

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"

    # The importer reads headers from row 3 and data from row 4, so place
    # cells explicitly rather than relying on append() row bookkeeping.
    headers = ["Folder", "Project No", "Project Description / Address",
               "Priority", "Project Status", "Action By", "Next Action",
               "Date Opened", "Target Close", "% Complete", "City", "Contact",
               "Company / Client", "Scope of Work", "Contract Value ($)",
               "Amount Paid ($)", "Outstanding Balance ($)", "COGS", "Profit",
               "Notes"]
    for i, name in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=name)

    def _write(row, job, name, status, pct):
        values = [None, job, name, None, status, None, None, None, None, pct,
                  None, None, None, None, 1000, 0, 1000, 0, 0, None]
        for i, value in enumerate(values, start=1):
            ws.cell(row=row, column=i, value=value)

    _write(4, "260201", "Done Job", "Completed", 0.01)
    _write(5, "260202", "Live Job", "Drafting", 0.30)

    trk.import_projects(db, wb)

    rows = {r["job_number"]: r for r in db.execute(
        "SELECT job_number, status, percent_complete FROM projects"
    ).fetchall()}
    assert rows["260201"]["percent_complete"] == 100.0
    assert rows["260201"]["status"] == "completed"
    # An in-flight project keeps whatever the tracker says.
    assert rows["260202"]["percent_complete"] == 30.0
    assert rows["260202"]["status"] == "drafting"
