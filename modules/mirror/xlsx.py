"""Portfolio _AUTO_portfolio_overview.xlsx renderer.

Pure function — takes a list of project dicts and returns workbook bytes.
The 22-column layout intentionally mirrors the legacy
`Project_Tracker_2026.xlsx` so users see a familiar shape.

Idempotency: openpyxl writes workbook timestamps and a "modified by"
property by default, which would defeat sha256 short-circuiting. We pin
`wb.properties.created/modified` to a fixed epoch and clear the user-
identifying fields so successive renders with identical input produce
byte-identical output.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Iterable, Mapping

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "Folder", "Project No", "Project Description", "Priority", "Project Status",
    "Action By", "Next Action", "Date Opened", "Target Close", "% Complete",
    "Age (Days)", "Contact", "Company / Client", "Scope of Work",
    "Contract Value ($)", "Amount Paid ($)", "Outstanding Balance ($)",
    "COGS", "Profit", "Process No.", "Notes", "Link",
]

# Banner row above the headers — A1 must read prominently as a warning.
BANNER_TEXT = "AUTO-GENERATED — edit in 6DE Platform"

# Fixed epoch for workbook metadata to keep openpyxl output deterministic.
_EPOCH = datetime(2024, 1, 1, 0, 0, 0)

_PRIORITY_FILLS = {
    "urgent": PatternFill("solid", fgColor="EF4444"),
    "high":   PatternFill("solid", fgColor="F59E0B"),
    "normal": PatternFill("solid", fgColor="FFFFFF"),
    "low":    PatternFill("solid", fgColor="22C55E"),
}

_STATUS_FILLS = {
    "active":         PatternFill("solid", fgColor="1FBA66"),
    "prospect":       PatternFill("solid", fgColor="F7B500"),
    "drafting":       PatternFill("solid", fgColor="3B82F6"),
    "ahj_permitting": PatternFill("solid", fgColor="F59E0B"),
    "inspection":     PatternFill("solid", fgColor="06B6D4"),
    "revisions":      PatternFill("solid", fgColor="EF4444"),
    "on_hold":        PatternFill("solid", fgColor="A85FFF"),
    "completed":      PatternFill("solid", fgColor="9CA3AF"),
    "cancelled":      PatternFill("solid", fgColor="6B7280"),
    "archived":       PatternFill("solid", fgColor="374151"),
}


def _platform_link(project: Mapping, base_url: str) -> str:
    job = project.get("job_number") or project.get("id") or ""
    return f"{base_url.rstrip('/')}/Projects?job={job}"


def _age_days(start_date: str | None, today: date | None = None) -> int | None:
    if not start_date:
        return None
    try:
        start = date.fromisoformat(str(start_date)[:10])
    except ValueError:
        return None
    ref = today or date.today()
    return (ref - start).days


def _truncate(value: str | None, limit: int = 200) -> str:
    if not value:
        return ""
    s = str(value)
    return s if len(s) <= limit else s[: limit - 1] + "…"


def _project_row(p: Mapping, base_url: str, today: date | None = None) -> list:
    return [
        p.get("folder_path") or "",
        p.get("job_number") or "",
        p.get("name") or "",
        p.get("priority") or "",
        p.get("status") or "",
        p.get("action_by") or "",
        p.get("next_action") or "",
        str(p.get("start_date") or "")[:10],
        str(p.get("target_end_date") or "")[:10],
        float(p.get("percent_complete") or 0),
        _age_days(p.get("start_date"), today=today),
        p.get("contact_name") or "",
        p.get("client_name") or "",
        p.get("scope") or "",
        float(p.get("contract_value") or 0),
        float(p.get("amount_paid") or 0),
        float(p.get("outstanding_balance") or 0),
        float(p.get("cogs") or 0),
        float(p.get("profit") or 0),
        p.get("process_no") or "",
        _truncate(p.get("notes")),
        _platform_link(p, base_url),
    ]


def _apply_priority_fills(ws, first_data_row: int, last_data_row: int) -> None:
    col = HEADERS.index("Priority") + 1
    for row in range(first_data_row, last_data_row + 1):
        cell = ws.cell(row=row, column=col)
        fill = _PRIORITY_FILLS.get(str(cell.value or "").lower())
        if fill:
            cell.fill = fill


def _apply_status_fills(ws, first_data_row: int, last_data_row: int) -> None:
    col = HEADERS.index("Project Status") + 1
    for row in range(first_data_row, last_data_row + 1):
        cell = ws.cell(row=row, column=col)
        fill = _STATUS_FILLS.get(str(cell.value or "").lower())
        if fill:
            cell.fill = fill


def _apply_percent_databar(ws, first_data_row: int, last_data_row: int) -> None:
    col_letter = get_column_letter(HEADERS.index("% Complete") + 1)
    rng = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
    ws.conditional_formatting.add(
        rng,
        DataBarRule(
            start_type="num", start_value=0,
            end_type="num", end_value=100,
            color="1FBA66", showValue=True,
        ),
    )


def _apply_outstanding_completed_rule(ws, first_data_row: int, last_data_row: int) -> None:
    """Outstanding > 0 on a completed project is a red-text flag."""
    col_letter = get_column_letter(HEADERS.index("Outstanding Balance ($)") + 1)
    rng = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
    red_font = Font(color="EF4444", bold=True)
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="greaterThan", formula=["0"], font=red_font),
    )


def _write_projects_sheet(ws, projects: list[Mapping], base_url: str, today: date | None):
    ws.title = "Projects"

    # Banner row.
    ws.cell(row=1, column=1, value=BANNER_TEXT)
    banner = ws.cell(row=1, column=1)
    banner.font = Font(bold=True, color="EF4444", size=14)
    banner.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    # Header row.
    for i, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F3F4F6")

    # Data rows.
    first_data = 3
    for ridx, p in enumerate(projects, start=first_data):
        for cidx, val in enumerate(_project_row(p, base_url, today=today), start=1):
            ws.cell(row=ridx, column=cidx, value=val)

    last_data = first_data + len(projects) - 1

    if projects:
        _apply_priority_fills(ws, first_data, last_data)
        _apply_status_fills(ws, first_data, last_data)
        _apply_percent_databar(ws, first_data, last_data)
        _apply_outstanding_completed_rule(ws, first_data, last_data)

    # Freeze the header so scrolling stays oriented.
    ws.freeze_panes = "A3"


def _write_generated_sheet(
    ws, projects: list[Mapping], platform_version: str, today: date,
    skipped: list[tuple[str, str]],
):
    ws.title = "Generated"
    ws.cell(row=1, column=1, value="Field")
    ws.cell(row=1, column=2, value="Value")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)

    meta = [
        ("Generated on (UTC date)", today.isoformat()),
        ("Source", "6DE Platform"),
        ("Platform version", platform_version),
        ("Project count", len(projects)),
        ("Skipped count", len(skipped)),
    ]
    for i, (k, v) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

    if skipped:
        skip_start = len(meta) + 3
        ws.cell(row=skip_start, column=1, value="Skipped").font = Font(bold=True)
        ws.cell(row=skip_start, column=2, value="Reason").font = Font(bold=True)
        for i, (job, reason) in enumerate(skipped, start=skip_start + 1):
            ws.cell(row=i, column=1, value=job)
            ws.cell(row=i, column=2, value=reason)


def render_portfolio_overview(
    projects: Iterable[Mapping],
    *,
    base_url: str = "http://localhost:8501",
    platform_version: str = "v3.5",
    today: date | None = None,
    skipped: list[tuple[str, str]] | None = None,
) -> bytes:
    """Render the portfolio _AUTO_portfolio_overview.xlsx as bytes.

    ``today`` and ``platform_version`` are injected for deterministic tests
    and stable per-day output (required for sha256 short-circuiting).
    """
    today = today or date.today()
    projects_list = list(projects)
    skipped = skipped or []

    wb = Workbook()

    # Strip user-identifying + drift-prone metadata so the .xlsx bytes are
    # deterministic across hosts and successive runs on the same day.
    wb.properties.creator = "6DE Platform"
    wb.properties.lastModifiedBy = "6DE Platform"
    wb.properties.created = _EPOCH
    wb.properties.modified = _EPOCH
    wb.properties.lastPrinted = None
    wb.properties.title = "6DE Portfolio Overview"
    wb.properties.subject = "Auto-generated portfolio snapshot"
    wb.properties.description = (
        "Generated by 6DE Platform. Do not edit by hand — "
        "changes will be overwritten on next regeneration."
    )

    projects_ws = wb.active
    _write_projects_sheet(projects_ws, projects_list, base_url, today)

    generated_ws = wb.create_sheet("Generated")
    _write_generated_sheet(generated_ws, projects_list, platform_version, today, skipped)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
