"""Weekly timesheet .xlsx export — Excel-system parity.

Builds a workbook that matches the firm's hand-filled weekly timesheet
files (``Timesheet_<Monday>_to_<Sunday>.xlsx``) closely enough that the
downstream Master_Time_Log builder — which reads the ``Time_Log`` sheet
rows having both a Date and Hours — keeps working unchanged:

- ``Fee_Rates``      role display labels + std / OT (1.5x) rates
- ``Internal_Codes`` the internal (non-billable) code list from the DB
- ``Time_Log``       EXACT legacy layout: title rows 1-2, headers on
                     row 5, entries from row 6 as literal values
- ``Weekly_Summary`` B2 = Monday date + hours/revenue by role

The Excel files' Time_Analysis and Management_Quadrant sheets are NOT
reproduced — they are in-file conveniences, not data (see
docs/timesheets.md).
"""

from __future__ import annotations

import io
import sqlite3
from datetime import date, datetime, timedelta

from modules.timekeeping.crud import (
    AFTER_HOURS_MULTIPLIER,
    get_weekly_timesheet,
    list_fee_schedule,
    list_internal_codes,
)

# Role enum -> the display label used in the Excel system (Fee_Rates A2:A8
# and the Time_Log "Role" column). Order matters: it is the Fee_Rates /
# Weekly_Summary row order.
ROLE_DISPLAY = {
    "principal": "Principal",
    "expert_consultant": "Expert Consulting",
    "professional_engineer": "Professional Engineer",
    "field_inspector": "Field Inspector",
    "engineering_technician": "Engineering Technician",
    "cad_drafter": "CAD Drafter",
    "admin": "Administrative/Clerical",
}

# Reverse map for the importer (display label -> enum).
ROLE_FROM_DISPLAY = {v: k for k, v in ROLE_DISPLAY.items()}

TIME_LOG_HEADERS = [
    "Date", "Project #", "Client / Project Name", "Task Description",
    "Role", "OT?", "Hours", "Rate ($/hr)", "Line Total ($)", "Mileage",
]

TIME_LOG_HEADER_ROW = 5  # headers on row 5, entries from row 6


def timesheet_filename(monday: date) -> str:
    """``Timesheet_<Monday>_to_<Sunday>.xlsx``"""
    sunday = monday + timedelta(days=6)
    return f"Timesheet_{monday.isoformat()}_to_{sunday.isoformat()}.xlsx"


def _current_rates(conn: sqlite3.Connection) -> dict[str, float]:
    """Latest std rate per role from fee_schedule (greatest effective_date)."""
    rates: dict[str, float] = {}
    for row in list_fee_schedule(conn):
        # list_fee_schedule orders by role, effective_date DESC — first
        # row seen per role is the current one.
        if row["role"] not in rates:
            rates[row["role"]] = float(row["hourly_rate"])
    return rates


def build_timesheet_workbook(
    conn: sqlite3.Connection,
    employee_id: int,
    monday: date,
) -> bytes:
    """Build the weekly timesheet workbook and return it as xlsx bytes."""
    from openpyxl import Workbook

    wb = Workbook()

    # -- Fee_Rates ---------------------------------------------------------
    ws = wb.active
    ws.title = "Fee_Rates"
    ws.append(["Role", "Std Rate ($/hr)", "OT Rate ($/hr)"])
    rates = _current_rates(conn)
    for role, label in ROLE_DISPLAY.items():
        std = rates.get(role)
        ot = round(std * AFTER_HOURS_MULTIPLIER, 2) if std is not None else None
        ws.append([label, std, ot])

    # -- Internal_Codes ----------------------------------------------------
    ws = wb.create_sheet("Internal_Codes")
    ws.append(["6th Degree Engineering - Internal Project Codes"])
    ws.append(['Use these codes in the Time_Log "Project #" column for '
               "non-billable / internal work"])
    ws.append(["Project #", "Category", "Description"])
    for code in list_internal_codes(conn):
        ws.append([code["code"], code["category"], code["description"]])

    # -- Time_Log ----------------------------------------------------------
    ws = wb.create_sheet("Time_Log")
    ws.cell(row=1, column=1, value="6th Degree Engineering - Billable Time Log")
    ws.cell(row=2, column=1, value=(
        "Exported from the 6DE platform. Literal values - no formulas."
    ))
    for col, header in enumerate(TIME_LOG_HEADERS, start=1):
        ws.cell(row=TIME_LOG_HEADER_ROW, column=col, value=header)

    entries = get_weekly_timesheet(conn, employee_id, monday.isoformat())
    role_totals: dict[str, tuple[float, float]] = {
        role: (0.0, 0.0) for role in ROLE_DISPLAY
    }
    out_row = TIME_LOG_HEADER_ROW + 1
    for e in entries:
        multiplier = float(e["multiplier"] or 1.0)
        rate = float(e["rate"])
        hours = float(e["hours"])
        eff_rate = round(rate * multiplier, 2)   # OT rows show the OT rate
        line_total = round(hours * rate * multiplier, 2)
        is_internal = e["internal_code"] is not None
        project_no = e["internal_code"] if is_internal else e["job_number"]
        client_name = (
            e["internal_description"] if is_internal else e["project_name"]
        )
        entry_date = datetime.strptime(str(e["entry_date"])[:10], "%Y-%m-%d")
        row = [
            entry_date,                             # A Date — real date cell
            project_no,                             # B Project #
            client_name,                            # C Client / Project Name
            e["description"],                       # D Task Description
            ROLE_DISPLAY.get(e["role"], e["role"]),  # E Role (display label)
            "Y" if multiplier > 1 else "N",         # F OT?
            hours,                                  # G Hours
            eff_rate,                               # H Rate ($/hr)
            line_total,                             # I Line Total ($)
            None,                                   # J Mileage (not tracked)
        ]
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=out_row, column=col, value=val)
            if col == 1:
                cell.number_format = "mm/dd/yyyy"
        out_row += 1

        h, rev = role_totals.get(e["role"], (0.0, 0.0))
        role_totals[e["role"]] = (h + hours, rev + line_total)

    # -- Weekly_Summary ------------------------------------------------------
    ws = wb.create_sheet("Weekly_Summary")
    ws.cell(row=1, column=1, value="Weekly Summary Dashboard")
    ws.cell(row=2, column=1, value="Week Starting:")
    b2 = ws.cell(row=2, column=2, value=datetime.combine(monday, datetime.min.time()))
    b2.number_format = "mm/dd/yyyy"
    ws.cell(row=4, column=1, value="Hours & Revenue by Role")
    ws.append([])  # spacer safety not needed; we place explicitly below
    ws.cell(row=5, column=1, value="Role")
    ws.cell(row=5, column=2, value="Hours")
    ws.cell(row=5, column=3, value="Revenue ($)")
    r = 6
    total_h = total_rev = 0.0
    for role, label in ROLE_DISPLAY.items():
        h, rev = role_totals.get(role, (0.0, 0.0))
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=round(h, 2))
        ws.cell(row=r, column=3, value=round(rev, 2))
        total_h += h
        total_rev += rev
        r += 1
    ws.cell(row=r, column=1, value="TOTAL")
    ws.cell(row=r, column=2, value=round(total_h, 2))
    ws.cell(row=r, column=3, value=round(total_rev, 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
